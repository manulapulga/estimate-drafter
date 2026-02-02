import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.styles.colors import Color

def generate_bill_excel(
    selected_items,
    tender_mode,
    tender_percent,
    deductions=None
):
    # Default deductions if not provided
    if deductions is None:
        deductions = {
            'it_enabled': True,
            'it_rate': 1,  # 1% for Individuals
            'welfare_enabled': True,
            'gst_enabled': False,
            'dept_enabled': False,
            'dept_desc': "Departmental Deduction",
            'dept_amount': 0.0,
            'fine_enabled': False,
            'fine_desc': "Fine",
            'fine_amount': 0.0,
            'kseb_enabled': False,
            'kseb_amount': 0.0,
            # ADD NEW FIELDS:
            'other_charges_enabled': False,
            'other_charges_desc': "Other Charges",
            'other_charges_amount': 0.0,
            'other_deductions_enabled': False,
            'other_deductions_desc': "Other Deductions",
            'other_deductions_amount': 0.0
        }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill"

    # ------------------ STYLES ------------------
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(bold=True)
    summary_font = Font(bold=True)
    
    # Color fills
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # ------------------ DYNAMIC HEADER ------------------
    if tender_mode == "Below":
        agreed_title = f"Agreed Rate (Below {tender_percent}% PAC)"
    elif tender_mode == "Above":
        agreed_title = f"Agreed Rate (Above {tender_percent}% PAC)"
    else:
        agreed_title = "Agreed Rate (At PAC)"

    headers = [
        "Sl No",
        "Description",
        "Qty",
        "Unit",
        "Estimate Rate",
        agreed_title,
        "Amount"
    ]

    ws.append(headers)

    # Header formatting
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # ------------------ ITEMS ------------------
    row = 2
    sl_no = 1

    for item in selected_items:

        # ---- Subheading ----
        if item.get("Type") == "Subheading":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(row=row, column=1, value=item["Item"])
            cell.alignment = left_align
            cell.border = border
            row += 1
            continue

        ws.cell(row=row, column=1, value=sl_no)
        ws.cell(row=row, column=2, value=item["Item"])
        ws.cell(row=row, column=3, value=item["Quantity"])
        ws.cell(row=row, column=4, value=item["Item Unit"])
        ws.cell(row=row, column=5, value=item["Unit Price"])

        # ---- AGREED RATE (FORMULA, 2 decimals) ----
        if tender_mode == "Below":
            ws.cell(row=row, column=6).value = (
                f"=ROUND(E{row}*(100-{tender_percent})/100,2)"
            )
        elif tender_mode == "Above":
            ws.cell(row=row, column=6).value = (
                f"=ROUND(E{row}*(100+{tender_percent})/100,2)"
            )
        else:  # At
            ws.cell(row=row, column=6).value = f"=ROUND(E{row},2)"

        # ---- AMOUNT (2 decimals) ----
        ws.cell(row=row, column=7).value = f"=ROUND(C{row}*F{row},2)"

        sl_no += 1
        row += 1

    # ------------------ SUMMARY (0 DECIMALS) ------------------
    sub_total_row = row

    # Sub Total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Sub Total")
    ws.cell(row=row, column=7, value=f"=ROUND(SUM(G2:G{row-1}),0)")
    subtotal_excl_gst = f"G{sub_total_row}"  # Save reference for deductions
    
    # Apply right alignment and bold to title cell
    title_cell = ws.cell(row=row, column=1)
    title_cell.alignment = right_align
    title_cell.font = summary_font
    
    # Apply center alignment and bold to value cell
    value_cell = ws.cell(row=row, column=7)
    value_cell.alignment = center_align
    value_cell.font = summary_font
    
    row += 1

    # GST
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="GST @18%")
    ws.cell(row=row, column=7, value=f"=ROUND({subtotal_excl_gst}*0.18,0)")
    gst_row = row
    
    # Apply right alignment and bold to title cell
    title_cell = ws.cell(row=row, column=1)
    title_cell.alignment = right_align
    title_cell.font = summary_font
    
    # Apply center alignment and bold to value cell
    value_cell = ws.cell(row=row, column=7)
    value_cell.alignment = center_align
    value_cell.font = summary_font
    
    row += 1

    # KSEB Charges (if applicable) - ADDED TO GRAND TOTAL, NOT A DEDUCTION
    kseb_amount = 0
    if 'kseb_enabled' in deductions and deductions['kseb_enabled']:
        kseb_amount = deductions.get('kseb_amount', 0)
        if kseb_amount > 0:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value="KSEB Charges")
            ws.cell(row=row, column=7, value=f"=ROUND({kseb_amount},0)")
            
            # Apply right alignment and bold to title cell
            title_cell = ws.cell(row=row, column=1)
            title_cell.alignment = right_align
            title_cell.font = summary_font
            
            # Apply center alignment and bold to value cell
            value_cell = ws.cell(row=row, column=7)
            value_cell.alignment = center_align
            value_cell.font = summary_font
            
            kseb_row = row
            row += 1

    # Other Charges (if applicable) - ADDED TO GRAND TOTAL, NOT A DEDUCTION
    other_charges_amount = 0
    if 'other_charges_enabled' in deductions and deductions['other_charges_enabled']:
        other_charges_amount = deductions.get('other_charges_amount', 0)
        if other_charges_amount > 0:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=deductions.get('other_charges_desc', 'Other Charges'))
            ws.cell(row=row, column=7, value=f"=ROUND({other_charges_amount},0)")
            
            # Apply right alignment and bold to title cell
            title_cell = ws.cell(row=row, column=1)
            title_cell.alignment = right_align
            title_cell.font = summary_font
            
            # Apply center alignment and bold to value cell
            value_cell = ws.cell(row=row, column=7)
            value_cell.alignment = center_align
            value_cell.font = summary_font
            
            other_charges_row = row
            row += 1

    # Grand Total (update to include KSEB and Other Charges if applicable)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Grand Total")
    
    # Build the Grand Total formula dynamically
    formula_parts = [f"{subtotal_excl_gst}", f"G{gst_row}"]
    
    if 'kseb_enabled' in deductions and deductions['kseb_enabled'] and kseb_amount > 0:
        formula_parts.append(f"G{kseb_row}")
    
    if 'other_charges_enabled' in deductions and deductions['other_charges_enabled'] and other_charges_amount > 0:
        formula_parts.append(f"G{other_charges_row}")
    
    formula = "=ROUND(" + "+".join(formula_parts) + ",0)"
    ws.cell(row=row, column=7, value=formula)
    
    # Apply green fill to both title and value cells for Grand Total
    for col in [1, 7]:
        cell = ws.cell(row=row, column=col)
        cell.fill = green_fill
    
    # Apply right alignment and bold to title cell
    title_cell = ws.cell(row=row, column=1)
    title_cell.alignment = right_align
    title_cell.font = summary_font
    
    # Apply center alignment and bold to value cell
    value_cell = ws.cell(row=row, column=7)
    value_cell.alignment = center_align
    value_cell.font = summary_font
    
    grand_total_row = row
    grand_total_ref = f"G{grand_total_row}"
    row += 1

    # ------------------ DEDUCTIONS SECTION ------------------
    # Start deductions from this row
    deduction_start_row = row
    deduction_rows_added = 0
    
    # Income Tax TDS
    if deductions['it_enabled']:
        it_rate = deductions['it_rate']
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"Deduction: {it_rate}% TDS Towards Income Tax")
        ws.cell(row=row, column=7, value=f"=ROUND({subtotal_excl_gst}*{it_rate/100},0)")
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        deduction_rows_added += 1
    
    # Workers Welfare Board
    if deductions['welfare_enabled']:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="Deduction: 1% Payment Towards Workers Welfare Board")
        ws.cell(row=row, column=7, value=f"=ROUND({subtotal_excl_gst}*0.01,0)")
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        deduction_rows_added += 1
    
    # GST TDS (2% of Subtotal excluding GST)
    if deductions['gst_enabled']:
        # Calculate GST TDS using the complex formula
        # Formula: IF(B28>250000, IF(MOD(ROUND(B28/50),2)=1, ROUND(B28/50)+1, ROUND(B28/50)), "NA")
        # Where B28 is subtotal_excl_gst
        gst_tds_formula = f'=IF(MOD(ROUND({subtotal_excl_gst}/50,0),2)=1, ROUND({subtotal_excl_gst}/50,0)+1, ROUND({subtotal_excl_gst}/50,0))'
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="Deduction: 2% TDS Towards GST")
        ws.cell(row=row, column=7, value=gst_tds_formula)
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        deduction_rows_added += 1
    
    # Departmental Deduction
    if deductions['dept_enabled'] and deductions['dept_amount'] > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"Deduction: {deductions['dept_desc']}")
        ws.cell(row=row, column=7, value=deductions['dept_amount'])
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        deduction_rows_added += 1
    
    # Fine
    if deductions['fine_enabled'] and deductions['fine_amount'] > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"Deduction: {deductions['fine_desc']}")
        ws.cell(row=row, column=7, value=deductions['fine_amount'])
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        deduction_rows_added += 1
    
    # Other Deductions (if applicable) - CONTRIBUTES TO DEDUCTIONS
    if deductions.get('other_deductions_enabled', False) and deductions.get('other_deductions_amount', 0) > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"Deduction: {deductions.get('other_deductions_desc', 'Other Deductions')}")
        ws.cell(row=row, column=7, value=deductions.get('other_deductions_amount', 0))
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        deduction_rows_added += 1
    
    # Calculate total deductions only if any deductions were added
    if deduction_rows_added > 0:
        # Total Deductions
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="Total Deductions")
        # Sum all deduction rows
        deduction_range = f"G{deduction_start_row}:G{row-1}"
        ws.cell(row=row, column=7, value=f"=SUM({deduction_range})")
        total_deductions_row = row
        total_deductions_ref = f"G{total_deductions_row}"
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1
        
        # Final Payment to Contractor
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="Final Payment to Contractor")
        ws.cell(row=row, column=7, value=f"={grand_total_ref}-{total_deductions_ref}")
        
        # Apply yellow fill to both title and value cells for final payment
        for col in [1, 7]:
            cell = ws.cell(row=row, column=col)
            cell.fill = yellow_fill
        
        # Apply right alignment and bold to title cell
        title_cell = ws.cell(row=row, column=1)
        title_cell.alignment = right_align
        title_cell.font = summary_font
        
        # Apply center alignment and bold to value cell
        value_cell = ws.cell(row=row, column=7)
        value_cell.alignment = center_align
        value_cell.font = summary_font
        
        row += 1

    # ------------------ FORMATTING ------------------
    for r in range(1, ws.max_row + 1):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = border

            # Apply formatting for regular rows (not summary rows)
            if r == 1:
                cell.alignment = center_align
            elif r >= 2:
                # Skip formatting for summary rows (already formatted)
                # Check if this is a regular row by checking if it's not in the summary section
                if r < sub_total_row:
                    if c == 2:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

    # ------------------ COLUMN WIDTHS ------------------
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 15

    # ------------------ OUTPUT ------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output