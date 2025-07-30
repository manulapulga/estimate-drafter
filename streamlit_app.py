import os
import io
import json
import base64
import math
import re
import pandas as pd
import streamlit as st
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from item_wizard import show_item_wizard
from decimal import Decimal, ROUND_HALF_UP, getcontext
from num2words import num2words
import firebase_admin
from firebase_admin import credentials, db
import time



# Set page config
st.set_page_config(
    page_title="Estimate Drafter",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Initialize Firebase once
if not firebase_admin._apps:
    try:
        # Convert st.secrets.AttrDict to a normal JSON-serializable dict
        firebase_dict = dict(st.secrets["firebase"])
        firebase_json = json.loads(json.dumps(firebase_dict))  # now it's safe for Certificate()

        # Initialize Firebase
        cred = credentials.Certificate(firebase_json)
        firebase_admin.initialize_app(cred, {
            'databaseURL': firebase_json["databaseURL"]
        })

        st.write("✅ Firebase Initialized")
    except Exception as e:
        st.error(f"❌ Firebase init failed: {e}")

def save_progress_to_firebase(username):
    try:
        # Convert session state to regular dict and remove unserializable items
        state_to_save = {
            k: v for k, v in st.session_state.items()
            if isinstance(v, (str, int, float, bool, list, dict, type(None)))
        }
        db.reference(f'progress/{username}').set(state_to_save)
        # Show temporary success message
        msg = st.empty()
        msg.success("✅ Progress saved successfully!")
        time.sleep(3)  # visible for 3 seconds
        msg.empty()
    except Exception as e:
        st.error(f"❌ Failed to save progress: {e}")

def load_progress_from_firebase(username):
    try:
        # Check if user has unsaved progress in session
        dirty = (
            st.session_state.get("selected_items") or
            st.session_state.get("file_name") or
            st.session_state.get("estimate_date") or
            st.session_state.get("work_desc") or
            st.session_state.get("head_note") or
            st.session_state.get("estimate_note")
        )

        if dirty:
            st.warning("⚠️ Please clear all data before loading progress.")
            return

        # Proceed with loading
        saved_state = db.reference(f'progress/{username}').get()
        if saved_state:
            # Restore basic fields immediately
            safe_keys = ['file_name', 'estimate_date', 'work_desc', 'head_note', 'estimate_note']
            for k in safe_keys:
                if k in saved_state:
                    st.session_state[k] = saved_state[k]

            # Restore items progressively
            if "selected_items" in saved_state:
                st.session_state["__pending_restore_items__"] = saved_state["selected_items"]
                st.session_state["__restore_index__"] = 0

            st.success("⏳ Restoring... please wait.")
            st.rerun()
        else:
            st.warning("⚠️ No saved progress found.")
    except Exception as e:
        st.error(f"❌ Failed to load progress: {e}")



getcontext().prec = 10  # Increase precision


# Hide default Streamlit sidebar navigation
st.markdown("""
    <style>
        [data-testid="stSidebarNav"] {
            display: none;
        }
    </style>
""", unsafe_allow_html=True)

# Auto-open wizard if flagged
if st.session_state.get("wizard_target_index") is not None:
    st.switch_page("pages/wizard.py")
    
st.markdown("""
    <style>
    /* Make all buttons stretch full width of their container */
    button[kind="primary"], button[kind="secondary"] {
        width: 100% !important;
        display: block !important;
    }
    .stButton > button {
        width: 100% !important;
    }
    </style>
""", unsafe_allow_html=True)


st.markdown("""
<style>
/* Extra-light green to blue diagonal background */
body, .stApp {
    background: linear-gradient(135deg, #f6fdf4 0%, #f4f9fd 100%) !important;
    background-attachment: fixed;
    background-repeat: no-repeat;
    background-size: cover;
}
</style>
""", unsafe_allow_html=True)



# Inject styles early
st.markdown("""
<style>
/* Pale green gradient for collapsed expanders */
div[data-testid="stExpander"]:not([open]) {
    background: linear-gradient(135deg, #f3fdf4 0%, #e3f8e5 50%, #d0f0d2 100%) !important;
    border: 1px solid #b2ddb3 !important;
    border-radius: 10px !important;
    margin-bottom: 14px;
    padding: 0 !important;
    color: #1e4023 !important;
}

/* Expander header styling */
div[data-testid="stExpander"]:not([open]) > summary {
    background-color: transparent !important;
    padding: 12px 18px !important;
    font-family: 'Segoe UI', 'Roboto', sans-serif;
    font-weight: 600;
    font-size: 16px;
    color: #1e4023 !important;
    text-transform: uppercase;
    letter-spacing: 0.4px;
    border-radius: 10px !important;
    border: none !important;
    cursor: pointer;
}

/* Hover effect — slightly deeper green gradient */
div[data-testid="stExpander"]:not([open]):hover {
    background: linear-gradient(135deg, #e7faea 0%, #d4f2d8 50%, #c2e8c5 100%) !important;
}
</style>
""", unsafe_allow_html=True)


# Load user credentials from Sheet 2 of Excel
@st.cache_data
def load_credentials(file_path):
    return pd.read_excel(file_path, sheet_name=1)  # Sheet 2 is index 1

def save_excel_to_firebase(username, filename, excel_io):
    def get_unique_filename(base_name):
        existing_files = list_user_files(username).keys()
        if base_name not in existing_files:
            return base_name
        i = 1
        while f"{base_name}_{i}" in existing_files:
            i += 1
        return f"{base_name}_{i}"

    # Ensure file pointer is at start
    excel_io.seek(0)
    encoded = base64.b64encode(excel_io.read()).decode()

    unique_filename = get_unique_filename(filename)
    db.reference(f'estimates/{username}/{unique_filename}').set(encoded)
    return unique_filename  # optional, useful for UI confirmation
    

def list_user_files(username):
    files = db.reference(f'estimates/{username}').get()
    return files if files else {}

def load_excel_from_firebase(username, filename):
    encoded = db.reference(f'estimates/{username}/{filename}').get()
    if encoded:
        return io.BytesIO(base64.b64decode(encoded))
    return None
    
def break_long_words(text, max_length=8):
    """
    Insert a space after every max_length characters in long unbroken words.
    """
    if not text:
        return text
    return ' '.join([
        word if len(word) <= max_length else ' '.join([word[i:i+max_length] for i in range(0, len(word), max_length)])
        for word in text.split()
    ])

def force_refresh():
    st.session_state._force_refresh = not st.session_state.get("_force_refresh", False)
    st.rerun()  # Immediately rerun app
  
def toggle_section(section_key):
    """Toggle a section and properly collapse all others"""
    # Get current state of the clicked section
    current_state = st.session_state.get(section_key, False)
    
    # List of all expandable section keys
    all_sections = [
        'show_dsr_options',
        'show_price_options',
        'show_dsr21basicrates_options',
        'show_priceapprovedmr_options',
        'show_costindex_options',
        'show_gwd_options',
        'show_templates_options',
        'show_pump_selector'
    ]
    
    # If we're opening this section (it was previously closed)
    if not current_state:
        # Close all other sections
        for key in all_sections:
            if key != section_key:
                st.session_state[key] = False
    
    # Toggle the current section
    st.session_state[section_key] = not current_state           
            
# Authentication function
def authenticate(username, password, credentials_df):
    user_row = credentials_df[credentials_df['username'] == username]
    if not user_row.empty:
        return user_row.iloc[0]['password'] == password
    return False

# Load main items data
@st.cache_data
def load_main_items(username):
    try:
        data = pd.read_excel("Data Base/items.xltm", sheet_name=username)
        return data['Item Name'].tolist(), data['Unit Price'].tolist(), data['Item Unit'].tolist(), data
    except Exception as e:
        st.error(f"Error loading main items data for {username}: {str(e)}")
        st.stop()
        
@st.cache_data
def load_templates():
    try:
        import os
        template_data = {}
        templates_dir = "Data Base/Global Templates"
        
        if os.path.exists(templates_dir) and os.path.isdir(templates_dir):
            # Get all Excel files in the Templates directory
            template_files = [f for f in os.listdir(templates_dir) 
                           if f.endswith(('.xlsx', '.xls')) and os.path.isfile(os.path.join(templates_dir, f))]
            
            for template_file in template_files:
                # Remove file extension for the sheet name
                sheet_name = os.path.splitext(template_file)[0]
                file_path = os.path.join(templates_dir, template_file)
                
                # Read each Excel file
                template_data[sheet_name] = pd.read_excel(file_path)
        
        if not template_data:
            st.warning("No template files found in the Templates directory")
            return {}
            
        return template_data
    except Exception as e:
        st.error(f"Error loading template data: {str(e)}")
        st.stop()
        
@st.cache_data
def load_local_templates(username):
    try:
        import os
        template_data = {}
        templates_dir = os.path.join("Data Base/Local Templates", username)
        
        if os.path.exists(templates_dir) and os.path.isdir(templates_dir):
            # Get all Excel files in the user's Local Templates directory
            template_files = [f for f in os.listdir(templates_dir) 
                           if f.endswith(('.xlsx', '.xls')) and os.path.isfile(os.path.join(templates_dir, f))]
            
            for template_file in template_files:
                # Remove file extension for the sheet name
                sheet_name = os.path.splitext(template_file)[0]
                file_path = os.path.join(templates_dir, template_file)
                
                # Read each Excel file
                template_data[sheet_name] = pd.read_excel(file_path)
        
        if not template_data:
            st.warning(f"No template files found in your local templates directory ({templates_dir})")
            return {}
            
        return template_data
    except Exception as e:
        st.error(f"Error loading local template data: {str(e)}")
        st.stop()
# Load wizard items data
@st.cache_data
def load_wizard_items(username):
    try:
        wizard_data = pd.read_excel("Data Base/items.xltm", sheet_name=username)
        return wizard_data
    except Exception as e:
        st.error(f"Error loading wizard items data for {username}: {str(e)}")
        st.stop()

# Login screen
import streamlit as st

# Load the banner image and encode it as base64
with open("Icons/banner.png", "rb") as image_file:
    encoded_banner = base64.b64encode(image_file.read()).decode()
    
    
with open("Icons/spec1.png", "rb") as image_file:
    encoded_logospec = base64.b64encode(image_file.read()).decode() 

def login_page(credentials_df):
    import base64

    # Load banner and logo
    with open("Icons/banner.png", "rb") as image_file:
        encoded_banner = base64.b64encode(image_file.read()).decode()
    with open("Icons/spec1.png", "rb") as image_file:
        encoded_logospec = base64.b64encode(image_file.read()).decode()

     # Inject CSS to remove top space
    st.markdown("""
        <style>
        html, body, .stApp {
            margin-top: 0px !important;
            padding-top: 0px !important;
        }
        [data-testid="stVerticalBlock"] {
            margin-top: 0px !important;
            padding-top: 0px !important;
        }
        .block-container {
            padding-top: 0rem !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <style>
        html, body, .stApp {
            background-color: #e3f2fd !important;  /* soft blue background */
            background-attachment: fixed;
            background-size: cover;
        }
        .st-emotion-cache-46x3d8 {
            width: 390px;
            position: relative;
            display: flex;
            flex: 1 1 0%;
            flex-direction: column;
            gap: 0rem;
        }
        .st-emotion-cache-n3bfp {
            display: inline-flex;
            -webkit-box-align: center;
            align-items: center;
            -webkit-box-pack: center;
            justify-content: center;
            font-weight: 400;
            padding: 0.25rem 0.75rem;
            border-radius: 0.5rem;
            min-height: 2.5rem;
            margin: 0px;
            line-height: 1.6;
            text-transform: none;
            font-size: inherit;
            font-family: inherit;
            color: inherit;
            width: auto;
            cursor: pointer;
            user-select: none;
            background-color: #65d1ea;
            border: 1px solid rgba(0, 0, 0, 0.2);
        }
    
        /* Add soft shadow illusion around center "card" */
        .block-container {
            max-width: 400px;
            margin-top: 80px;
            padding: 5px;
            box-shadow: 0 0 25px rgba(0,0,0,0.15);
            border-radius: 16px;
            background-color: rgba(255, 255, 255, 0.65); /* optional light tint */
            padding-top: 0px !important;
        }
    
        /* Optional: style input fields to pop more */
        input, textarea {
            background-color: rgba(255, 255, 255, 0.9) !important;
            border-radius: 8px !important;
            padding: 10px !important;
            border: 1px solid #b3cce6 !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Heading + logo section
    st.markdown(f"""
        <div style='text-align: center; margin-top: 0px; margin-bottom: 0px;'>
            <img src='data:image/png;base64,{encoded_logospec}' 
                 style='width: auto; max-height: 60px; object-fit: contain; border-radius: 8px; margin-bottom: 0px; margin-top: 25px;' />
            <h2 style='
                font-family: "Merriweather", serif;
                font-size: 20px;
                font-weight: 500;
                color: #1a4c74;
                margin-bottom: 50px;
            '>
                Standardised Project Estimation Console
            </h2>
        </div>
    """, unsafe_allow_html=True)

    # Center login inputs
    col1, col2, col3 = st.columns([1, 3, 1])

    with col2:
        username_input = st.text_input("", key="username_input", placeholder="Username", label_visibility="collapsed")
        password_input = st.text_input("", type="password", key="password_input", placeholder="Password", label_visibility="collapsed")

        if st.button("Login"):
            if authenticate(username_input, password_input, credentials_df):
                st.session_state.logged_in_username = username_input
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Invalid username or password")
                
    # Heading + logo section
    st.markdown(f"""
        <div style='text-align: center; margin-top: 0px; margin-bottom: 40px;'>
            <p style='
                font-family: "Segoe UI", sans-serif;
                font-size: 13px;
                color: #39779a;
                margin-top: 20px;
            '>
                Powered by DSR 2021
            </p>
        </div>
    """, unsafe_allow_html=True)


def set_rounding_option(option):
    """Handle mutually exclusive rounding options"""
    if option == 'manual':
        st.session_state.edit_final_total = True
        st.session_state.rounding_option = None
    else:
        st.session_state.edit_final_total = False
        st.session_state.rounding_option = option
    
    # Uncheck all other options
    for opt in ['1000', '100', 'none', 'manual']:
        if opt != option:
            key = f"{opt}_cb" if opt != 'manual' else "manual_edit_cb"
            if key in st.session_state:
                st.session_state[key] = False
    
    # Force immediate update
    st.rerun()    
# Main app
def main_app():
    # Load data
    username = st.session_state.logged_in_username
    
    
    
    # --- Step 1: Start progressive item restoration ---
    if "__pending_restore_items__" in st.session_state:
        items = st.session_state["__pending_restore_items__"]
        idx = st.session_state.get("__restore_index__", 0)
    
        if idx < len(items):
            if "selected_items" not in st.session_state:
                st.session_state.selected_items = []
    
            st.session_state.selected_items.append(items[idx])
            st.session_state["__restore_index__"] = idx + 1
            st.rerun()
        else:
            # Cleanup after all items are restored
            del st.session_state["__pending_restore_items__"]
            del st.session_state["__restore_index__"]




    
    item_names, unit_prices, item_units, data = load_main_items(username)
    wizard_data = load_wizard_items(username)
        
    # Pre-load work_desc, head_note, estimate_note, file_name, and estimate_date from uploaded_excel_data if available
    if 'uploaded_excel_data' in st.session_state:
        preload = st.session_state.uploaded_excel_data
        st.session_state.work_desc = preload.get("work_desc", "")
        st.session_state.head_note = preload.get("head_note", "")
        st.session_state.estimate_note = preload.get("estimate_note", "")
        st.session_state.file_name = preload.get("file_name", "")
        st.session_state.estimate_date = preload.get("estimate_date", None)
        del st.session_state.uploaded_excel_data  # Ensure one-time application

    # Handle clear all functionality using new query_params API
    if st.query_params.get("clear_all", "") == "true":
        st.session_state.work_desc = st.query_params.get("work_desc", "")
        st.session_state.head_note = st.query_params.get("head_note", "")
        st.session_state.estimate_note = st.query_params.get("estimate_note", "")
        st.session_state.file_name = ""
        st.session_state.estimate_date = None
        st.query_params.clear()  # Clear the query params after use
    # UI for Estimate Drafting with updated styles
    username = st.session_state.logged_in_username
    # Get the index value for the logged-in user
    user_row = credentials_df[credentials_df['username'] == username]
    cost_index = f"{user_row.iloc[0]['index'] + 1:.4f}" if not user_row.empty and 'index' in user_row.columns else "N/A"
    
    with open("Icons/spec1.png", "rb") as image_file:
        encoded_logo2spec = base64.b64encode(image_file.read()).decode()
    
    # Inject CSS to remove top space
    st.markdown("""
        <style>
        html, body, .stApp {
            margin-top: 0 !important;
            padding-top: 0 !important;
        }
        [data-testid="stVerticalBlock"] {
            margin-top: 0 !important;
            padding-top: 0 !important;
        }
        .block-container {
            padding-top: 1rem !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown(f"""
        <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@500&display=swap" rel="stylesheet">
        <style>
            @keyframes gradientSlide {{
                0% {{ background-position: 0% 50%; }}
                50% {{ background-position: 100% 50%; }}
                100% {{ background-position: 0% 50%; }}
            }}
    
            .ribbon-bar {{
                background: linear-gradient(270deg, #d0f0ff, #aee8f8, #d6d9f8, #c0e2ff);
                background-size: 600% 600%;
                animation: gradientSlide 10s ease infinite;
                color: #003344;
                display: flex;
                align-items: center;
                padding: 14px 24px;
                border-top-left-radius: 0px;
                border-top-right-radius: 0px;
                border-bottom-left-radius: 18px;
                border-bottom-right-radius: 18px;
                font-family: "Poppins", sans-serif;
                font-weight: 500;
                font-size: 17px;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                margin-bottom: 16px;
                letter-spacing: 0.4px;
            }}
    
            .ribbon-left,
            .ribbon-right {{
                flex: 1;
                text-align: center;
                font-size: 17px;
                font-weight: 500;
                color: #003344;
                white-space: nowrap;
            }}
    
            .ribbon-center {{
                flex: 2;
                text-align: center;
                font-size: 20px;
                font-weight: 600;
                letter-spacing: 1px;
                color: #003344;
                white-space: nowrap;
                display: flex;
                justify-content: center;
                align-items: center;
                gap: 10px;
            }}
    
            .ribbon-center img {{
                height: 28px;
                vertical-align: middle;
            }}
        </style>
    
        <div class="ribbon-bar">
            <div class="ribbon-left">Logged in as: {username}</div>
            <div class="ribbon-center">
                <img src="data:image/png;base64,{encoded_logo2spec}" alt="logo">
            </div>
            <div class="ribbon-right">Cost Index: {cost_index}</div>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <style>
            .stTextInput input {
                font-size: 100%;
                color: #1e81b0;
            }
            button {
                height: 50% !important;
            }
            .stSelectbox select {
                font-size: 18px;
            }
            .wizard-btn {
                background-color: #4CAF50 !important;
                color: white !important;
            }
            .estimate-item {
                margin-bottom: 1rem;
                border-radius: 0.5rem;
                background-color: #f8f9fa;
            }
            .item-actions {
                margin-top: 0.5rem;
            }
            .subheading-expander > .streamlit-expanderHeader {
                font-weight: bold !important;
            }
            .wizard-container {
                border: 1px solid #ddd;
                border-radius: 0.5rem;
                padding: 1rem;
                margin: 1rem 0;
                background-color: #f8f9fa;
            }
            .section-cancel-btn {
                background-color: #f44336 !important;
                color: white !important;
            }
            .other-item-btn {
                background-color: #FFA500 !important;
                color: white !important;
            }
            /* Make textarea more interactive */
            textarea[data-baseweb="textarea"] {
                min-height: 100px;
                resize: vertical;
                white-space: pre-wrap;
            }
            
            /* Ensure paste works in textarea */
            .stTextArea textarea {
                -webkit-user-select: text !important;
                user-select: text !important;
            }
            .stTextArea textarea {
                font-size: 16px !important;
                line-height: 1.5 !important;
                padding: 10px !important;
                min-height: 100px !important;
                resize: vertical !important;
            }
            /* Remove +/- buttons from number input */
            input[type=number]::-webkit-inner-spin-button, 
            input[type=number]::-webkit-outer-spin-button { 
                -webkit-appearance: none;
                margin: 0; 
            }
            input[type=number] {
                -moz-appearance: textfield;
            }
            
            /* Narrow number input */
            .narrow-number-input {
                width: 150px !important;
            }
            /* Apply to all number inputs */
            div[data-baseweb="input"] input {
                width: 150px !important;
            }
            /* Style for rounding option checkboxes */
            div[data-testid="stHorizontalBlock"] > div {
                padding: 0 5px;
            }
            div[data-testid="stHorizontalBlock"] label {
                font-size: 14px !important;
                white-space: nowrap;
            }
            div[data-testid="stHorizontalBlock"] .stCheckbox {
                margin-bottom: 0 !important;
            }
        </style>
    """, unsafe_allow_html=True)
    # Add Title row with File Name (left), Heading (center), Date (right), and Clear button
    st.markdown("""
        <style>
        .thin-banner {
            background: #5d89a3;
            padding: 2px 0px !important;
            border-top-left-radius: 13px;
            border-top-right-radius: 13px;
            border-bottom-left-radius: 0px;
            border-bottom-right-radius: 0px;
            box-shadow: 0 1px 4px rgba(0, 0, 0, 0.05);
            margin: 0px 0 6px 0;
            text-align: center;
            height: 30px;
            display: flex;
            justify-content: center;
            align-items: center;
        }
        .thin-banner h4 {
            color: white;
            font-size: 20px;
            font-weight: 600;
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1;
        }
        </style>
        <div class="thin-banner">
            <h4>ADD DETAILS</h4>
        </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns([1, 2, 2, 2])
    import re
    with col1:
        st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)
    
        file_name = st.text_input(
            label="File No",
            value=st.session_state.get("file_name", ""),
            key="file_name_input",
            label_visibility="collapsed",
            placeholder="File No"
        )
        st.session_state.file_name = file_name  # No sanitization
        
        # Make sure session state has a default key
        if "estimate_date" not in st.session_state:
            st.session_state.estimate_date = ""
        st.session_state.estimate_date = st.text_input(
            label="Estimate Date",
            value=st.session_state.estimate_date,
            key="estimate_date_input_text",
            label_visibility="collapsed",
            placeholder="DD/MM/YYYY"
        )
    with col2:  
        import re

        estimate_heading = st.text_area(
            "", 
            placeholder="Enter work description",
            value=st.session_state.get("work_desc", ""),
            key="work_desc_raw",
            height=100
        )
        st.session_state.work_desc = estimate_heading
        
    with col3:    
        if 'head_note' not in st.session_state:  # Add this line
            st.session_state.head_note = ""  
        # Add this section for head note
        head_note_container = st.container()
        with head_note_container:
            st.session_state.head_note = st.text_area(
                "",
                value=st.session_state.head_note,
                key="head_note_input",
                height=100,
                max_chars=2000,
                placeholder="Enter Head Note"
            )
    with col4:    
        if 'estimate_note' not in st.session_state:  # Add this line
            st.session_state.estimate_note = ""  
        # Add this section for head note
        estimate_note_container = st.container()
        with estimate_note_container:
            st.session_state.estimate_note = st.text_area(
                "",
                value=st.session_state.estimate_note,
                key="estimate_note_input",
                height=100,
                max_chars=2000,
                placeholder="Enter Foot Note"
            )
    
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    with col2: 
        if st.button("🗑 Clear Items", use_container_width=True):
            st.session_state.selected_items = []
            st.session_state.item_count = 0
            st.session_state.adding_subheading = False
            st.session_state.show_wizard = False
            st.session_state.show_add_item = False
            st.session_state.show_add_other = False
            st.rerun()
    with col3:
        if st.button("💾 Save Progress", use_container_width=True):
            save_progress_to_firebase(st.session_state.logged_in_username)
    with col4:
        if st.button("📂 Load Progress", use_container_width=True):
            load_progress_from_firebase(st.session_state.logged_in_username)
    with col1:        
        if st.button("🗑 Clear Details", use_container_width=True):
            st.session_state.file_name = ""
            st.session_state.estimate_date = ""
            st.session_state.work_desc = ""
            st.session_state.head_note = ""
            st.session_state.estimate_note = ""
            st.rerun()
        
    st.markdown("""
        <style>
        .thin-banner {
            background: #5d89a3;
            padding: 2px 0px !important;
            border-top-left-radius: 13px;
            border-top-right-radius: 13px;
            border-bottom-left-radius: 0px;
            border-bottom-right-radius: 0px;
            box-shadow: 0 1px 4px rgba(0, 0, 0, 0.05);
            margin: 0px 0 6px 0;
            text-align: center;
            height: 30px;
            display: flex;
            justify-content: center;
            align-items: center;
        }
        .thin-banner h4 {
            color: white;
            font-size: 20px;
            font-weight: 600;
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1;
        }
        </style>
        <div class="thin-banner">
            <h4>ADD ITEMS</h4>
        </div>
    """, unsafe_allow_html=True)
    st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
    # Initialize session state
    if 'selected_items' not in st.session_state:
        st.session_state.selected_items = []
    if 'item_count' not in st.session_state:
        st.session_state.item_count = 0
    if 'adding_subheading' not in st.session_state:
        st.session_state.adding_subheading = False
    if 'show_wizard' not in st.session_state:
        st.session_state.show_wizard = False
    if 'show_add_item' not in st.session_state:
        st.session_state.show_add_item = False
    if 'show_add_other' not in st.session_state:
        st.session_state.show_add_other = False
    if 'wizard_item_added' not in st.session_state:
        st.session_state.wizard_item_added = False
    if 'show_templates' not in st.session_state:
        st.session_state.show_templates = False    
    if 'show_upload' not in st.session_state:
        st.session_state.show_upload = False
    if 'show_local_templates' not in st.session_state:
        st.session_state.show_local_templates = False
    # Add this with your other session state initializations
    if 'show_preview' not in st.session_state:
        st.session_state.show_preview = False
    if 'manual_final_total' not in st.session_state:
        st.session_state.manual_final_total = None
    if 'edit_final_total' not in st.session_state:
        st.session_state.edit_final_total = False
    if 'signature_height' not in st.session_state:
        st.session_state.signature_height = 20  # or whatever default you want
    if 'signature_block_height' not in st.session_state:
        st.session_state.signature_block_height = 30 
    if 'signature_block_height' not in st.session_state:
        st.session_state.signature_block_height = 35
    if 'bottom_margin' not in st.session_state:
        st.session_state.bottom_margin = 20  # or any appropriate default value in mm    
    # Functions
    def update_all_items():
        selected_items = st.session_state.selected_items.copy()
        updated_count = 0
        
        for idx, item in enumerate(selected_items):
            if item.get('Type') == 'Subheading':
                # Handle subheading updates
                new_heading = st.session_state.get(f"edit_subheading_{idx}", item['Item'])
                if new_heading.strip() and new_heading != item['Item']:
                    selected_items[idx]['Item'] = new_heading.strip()
                    updated_count += 1
                continue
                
            if item.get('Type') == 'Other':
                # Handle 'Other' type items
                new_desc = st.session_state.get(f"other_desc_{idx}", item['Item'])
                new_qty = st.session_state.get(f"other_qty_{idx}", str(item['Quantity']))
                new_unit = st.session_state.get(f"other_unit_{idx}", item.get('Item Unit', ''))
                new_rate = st.session_state.get(f"other_rate_{idx}", str(item.get('Unit Price', 0)))
                new_gst = st.session_state.get(f"other_gst_{idx}", item.get('GST_Applicable', False))
                new_remark = break_long_words(st.session_state.get(f'remark_input_other_{idx}', item.get('Quantity_Remarks', '')))
                
                try:
                    qty = float(new_qty) if new_qty else 0
                    rate = float(new_rate) if new_rate else 0
                    if new_desc.strip():
                        selected_items[idx] = {
                            'Item': new_desc.strip(),
                            'Quantity': qty,
                            'Unit Price': rate,
                            'Item Unit': new_unit,
                            'Cost': qty * rate,
                            'Type': 'Other',
                            'GST_Applicable': new_gst,
                            'Quantity_Remarks': new_remark  # Apply break_long_words here
                        }
                        updated_count += 1
                except ValueError:
                    pass
            else:
                # Handle standard items
                item_name = st.session_state.get(f"edit_item_{idx}", '')
                quantity = st.session_state.get(f"edit_qty_{idx}", str(item['Quantity']))
                gst_applicable = st.session_state.get(f"edit_standard_gst_{idx}", item.get('GST_Applicable', True))
                new_remark = break_long_words(st.session_state.get(f'qty_remark_{idx}', item.get('Quantity_Remarks', '')))
                
                if item_name and quantity:
                    try:
                        quantity = float(quantity)
                        if quantity > 0:
                            item_data = data[data['Item Name'] == item_name]
                            if not item_data.empty:
                                item_data = item_data.iloc[0]
                                selected_items[idx] = {
                                    'Item': item_name,
                                    'Quantity': quantity,
                                    'Unit Price': item_data['Unit Price'],
                                    'Item Unit': item_data['Item Unit'],
                                    'Cost': quantity * item_data['Unit Price'],
                                    'Type': 'Standard',
                                    'GST_Applicable': gst_applicable,
                                    'Quantity_Remarks': new_remark  # Apply break_long_words here
                                }
                                updated_count += 1
                    except ValueError:
                        pass
        
        st.session_state.selected_items = selected_items
        return updated_count
    def add_item():
        st.session_state.item_count += 1

    def remove_item(index):
        st.session_state.selected_items.pop(index)
        st.session_state.item_count = max(0, st.session_state.item_count - 1)
        st.rerun()


    def calculate_totals(manual_unforeseen=None):
        """Calculate totals with manual unforeseen input
        Defaults to 2.5% of (total_cost + gst) if not provided
        Final total is rounded to next 100
        """
        selected_items = st.session_state.selected_items
        
        # Calculate total cost (excluding Subheadings)
        total_cost = sum(
            Decimal(str(item['Cost']))
            for item in selected_items
            if item.get('Type') != 'Subheading'
        ).to_integral_value(rounding=ROUND_HALF_UP)
        
        # Taxable amount (only GST-applicable items)
        taxable_amount = sum(
            Decimal(str(item['Cost']))
            for item in selected_items
            if item.get('Type') != 'Subheading' and item.get('GST_Applicable', True)
        )
        
        # GST at 18%, rounded to nearest rupee
        gst = (taxable_amount * Decimal('0.18')).to_integral_value(rounding=ROUND_HALF_UP)
        
        # Calculate base unforeseen (2.5% of total + gst)
        base_unforeseen = ((total_cost + gst) * Decimal('0.025')).to_integral_value(rounding=ROUND_HALF_UP)
        
        # Use manual unforeseen if provided, otherwise use base
        unforeseen = manual_unforeseen if manual_unforeseen is not None else base_unforeseen
        
        # Ensure unforeseen doesn't exceed 2.5% of total + gst
        unforeseen = min(unforeseen, base_unforeseen)
        
        # Calculate final total and round up to next 100
        # Rounding logic
        if st.session_state.edit_final_total and st.session_state.manual_final_total is not None:
            final_total = st.session_state.manual_final_total
        else:
            final_total = total_cost + gst + unforeseen
        
            # Apply rounding
            rounding = st.session_state.get('rounding_option', '1000')  # default
            if rounding == 'none':
                final_total = int(final_total)
            elif rounding == '100':
                final_total = math.ceil(final_total / 100) * 100
            elif rounding == '1000':
                final_total = math.ceil(final_total / 1000) * 1000
        
        
        return int(total_cost), int(gst), int(unforeseen), int(final_total)

    def move_item_up(index):
        if index > 0:
            st.session_state.selected_items[index], st.session_state.selected_items[index - 1] = \
                st.session_state.selected_items[index - 1], st.session_state.selected_items[index]
            st.rerun()
    
    def move_item_down(index):
        if index < len(st.session_state.selected_items) - 1:
            st.session_state.selected_items[index], st.session_state.selected_items[index + 1] = \
                st.session_state.selected_items[index + 1], st.session_state.selected_items[index]
            st.rerun()
    def move_item_to_position(current_index, new_position):
        selected_items = st.session_state.selected_items
        if 0 <= new_position < len(selected_items):
            item = selected_items.pop(current_index)
            selected_items.insert(new_position, item)
            st.rerun()
    def handle_item_selection(selected_item):
        # Check if we're editing an existing item
        if 'show_wizard_for_edit' in st.session_state and st.session_state.show_wizard_for_edit is not None:
            edit_idx = st.session_state.show_wizard_for_edit
            
            # Find the item in wizard data first
            wizard_item = wizard_data[wizard_data['Item Name'] == selected_item].iloc[0]
            
            # Try to find matching item in main data for unit price and unit
            main_item = data[data['Item Name'] == selected_item]
            
            if not main_item.empty:
                main_item = main_item.iloc[0]
                unit_price = main_item['Unit Price']
                unit = main_item['Item Unit']
            else:
                # Use wizard data if not found in main data
                unit_price = wizard_item['Unit Price']
                unit = wizard_item['Item Unit']
            
            # Preserve the existing quantity and remarks
            existing_item = st.session_state.selected_items[edit_idx]
            existing_quantity = existing_item['Quantity']
            existing_remarks = existing_item.get('Quantity_Remarks', '')
            
            # Update the item
            st.session_state.selected_items[edit_idx] = {
                'Item': selected_item,
                'Quantity': existing_quantity,
                'Unit Price': unit_price,
                'Item Unit': unit,
                'Cost': existing_quantity * unit_price,
                'Type': 'Standard',
                'GST_Applicable': True,
                'Quantity_Remarks': existing_remarks
            }
            
            # Close the wizard only when editing an item
            st.session_state.show_wizard = False
            st.session_state.show_wizard_for_edit = None
            st.rerun()
        else:
            # Original code for adding new items
            wizard_item = wizard_data[wizard_data['Item Name'] == selected_item].iloc[0]
            
            # Try to find matching item in main data for unit price and unit
            main_item = data[data['Item Name'] == selected_item]
            
            if not main_item.empty:
                main_item = main_item.iloc[0]
                unit_price = main_item['Unit Price']
                unit = main_item['Item Unit']
            else:
                # Use wizard data if not found in main data
                unit_price = wizard_item['Unit Price']
                unit = wizard_item['Item Unit']
            
            st.session_state.selected_items.append({
                'Item': selected_item,
                'Quantity': 1.0,
                'Unit Price': unit_price,
                'Item Unit': unit,
                'Cost': unit_price,
                'Type': 'Standard',
                'GST_Applicable': True,
                'Quantity_Remarks': ""
            })
            # Don't close the wizard when adding new items
            st.rerun()

    # Display added items and subheadings
    for idx, item in enumerate(st.session_state.selected_items):
        if item.get('Type') == 'Subheading':
            # Modified expander with controlled state
            expanded = st.session_state.get(f"expander_{idx}", False)
            # Add padding manually using Unicode spaces
            subheading_title = f"📂{item['Item']}📂"
            with st.expander(subheading_title, expanded=expanded):
                # Editable text input for subheading
                new_heading = st.text_input("Edit Subheading", value=item['Item'], key=f"edit_subheading_{idx}")
        
                # Update and Remove buttons
                col1, col2, col3, col4, col5, col6 = st.columns([1, 1, 1, 1, 0.5, 1])
                with col1:
                    if st.button("🔁 Update", key=f"update_sub_{idx}"):
                        if new_heading.strip():
                            st.session_state.selected_items[idx]['Item'] = new_heading.strip()
                            st.session_state[f"expander_{idx}"] = False  # Collapse the expander
                            st.rerun()
                with col2:
                    if st.button(f"❌ Remove", key=f"remove_sub_{idx}"):
                        remove_item(idx)
                with col3:    
                    if st.button("⬆️ Move Up", key=f"move_up_sub_{idx}"):
                        move_item_up(idx)
                with col4:
                    if st.button("⬇️ Move Down", key=f"move_down_sub_{idx}"):
                        move_item_down(idx)
                with col5:
                    new_pos = st.text_input(
                        "", 
                        key=f"move_input_{idx}", 
                        placeholder="Move To", 
                        label_visibility="collapsed"
                    )
                with col6:    
                    if st.button("Move", key=f"move_button_{idx}"):
                        try:
                            target_index = int(new_pos) - 1
                            if target_index == idx:
                                st.info("Item already at that position.")
                            elif 0 <= target_index < len(st.session_state.selected_items):
                                move_item_to_position(idx, target_index)
                            else:
                                st.warning("Invalid position number.")
                        except ValueError:
                            st.error("Please enter a valid number.")
                
            continue


        item_type = item.get('Type', 'Standard')
        item_title = f"💧 **Item {idx + 1}: {item['Item']} (₹{item['Cost']:.2f})**"
        if item_type == 'Other':
            item_title += " [Other" + (" +GST" if item.get('GST_Applicable', False) else "") + "]"
        
        # Modified expander with controlled state
        expanded = st.session_state.get(f"expander_{idx}", False)
        with st.expander(item_title, expanded=expanded):
                    
            if item_type == 'Other':
                # Enhanced display for "Other" type items with editing capability
                col1, col2, col3 = st.columns([5, 1, 1])
                with col1:
                    # Editable item description
                    new_desc = st.text_area(
                        "Item Description",
                        placeholder="Enter Item Name",
                        value=item['Item'],
                        key=f"other_desc_{idx}",
                        label_visibility="collapsed",
                        height=150  # Adjust height as needed
                    )
                with col2:
                    # Editable quantity
                    new_qty = st.text_input(
                        "Quantity",
                        placeholder="Quantity",
                        value=f"{item['Quantity']}",
                        key=f"other_qty_{idx}",
                        label_visibility="collapsed"
                    )
                    # Editable unit
                    new_unit = st.text_input(
                        "Unit",
                        placeholder="Unit",
                        value=item.get('Item Unit', ''),
                        key=f"other_unit_{idx}",
                        label_visibility="collapsed"
                    )
                    # Editable unit rate
                    new_rate = st.text_input(
                        "Unit Rate (₹)",
                        placeholder="Unit Rate (₹)",
                        value=f"{item.get('Unit Price', 0):.2f}",
                        key=f"other_rate_{idx}",
                        label_visibility="collapsed"
                    )
                with col3:    
                    # Calculate and display total
                    try:
                        qty = float(new_qty) if new_qty else 0
                        rate = float(new_rate) if new_rate else 0
                        total = qty * rate
                        st.markdown(f"**Total: ₹{total:,.2f}**")
                    except ValueError:
                        st.warning("Please enter valid numbers for quantity and rate")
                        
                    # Editable GST checkbox
                    new_gst = st.checkbox(
                        "GST Applicable?", 
                        value=item.get('GST_Applicable', False),
                        key=f"other_gst_{idx}"
                    )
                
                    # Remark Section for 'Other' Items
                    remark = item.get('Quantity_Remarks', '')
                    button_label = "✏️ Edit Remark" if remark else "➕ Add Remark"
                    
                    if st.button(button_label, key=f"edit_remark_other_{idx}"):
                        st.session_state.selected_items[idx]['show_remark_input'] = True
                    
                    if remark and not item.get('show_remark_input', False):
                        st.info(f"📋 Quantity Remark: {remark}")
                    
                    if item.get('show_remark_input', False):
                        new_remark = st.text_input("Edit Remark", value=remark, key=f"remark_input_other_{idx}", max_chars=100)
                        if st.button("Save Remark", key=f"save_remark_other_{idx}"):
                            st.session_state.selected_items[idx]['Quantity_Remarks'] = break_long_words(new_remark)
                            st.session_state.selected_items[idx]['show_remark_input'] = False
                            st.rerun()
            
                # Action buttons
                col1, col2, col3, col4, col5, col6 = st.columns([1, 1, 1, 1, 0.5, 1])
                with col1:
                    if st.button(f"🔁 Update", key=f"update_other_{idx}"):
                        if new_desc and new_qty and new_unit and new_rate:
                            try:
                                qty = float(new_qty)
                                rate = float(new_rate)
                                if qty > 0 and rate >= 0:
                                    st.session_state.selected_items[idx] = {
                                        'Item': new_desc,
                                        'Quantity': qty,
                                        'Unit Price': rate,
                                        'Item Unit': new_unit,
                                        'Cost': qty * rate,
                                        'Type': 'Other',
                                        'GST_Applicable': new_gst,
                                        'Quantity_Remarks': item.get('Quantity_Remarks', ''),
                                        'show_remark_input': False
                                    }
                                    st.session_state[f"expander_{idx}"] = False
                                    st.rerun()
                            except ValueError:
                                st.error("Please enter valid numbers for quantity and rate")
                with col2:
                    if st.button(f"❌ Remove", key=f"remove_{idx}"):
                        remove_item(idx)
                        
                with col3:
                    if st.button("⬆️ Move Up", key=f"move_up_sub_{idx}"):
                        move_item_up(idx)
                with col4:
                    if st.button("⬇️ Move Down", key=f"move_down_sub_{idx}"):
                        move_item_down(idx)
                with col5:
                    new_pos = st.text_input(
                        "", 
                        key=f"move_input_{idx}", 
                        placeholder="Move To", 
                        label_visibility="collapsed"
                    )
                with col6:    
                    if st.button("Move", key=f"move_button_{idx}"):
                        try:
                            target_index = int(new_pos) - 1
                            if target_index == idx:
                                st.info("Item already at that position.")
                            elif 0 <= target_index < len(st.session_state.selected_items):
                                move_item_to_position(idx, target_index)
                            else:
                                st.warning("Invalid position number.")
                        except ValueError:
                            st.error("Please enter a valid number.")        
            else:
                # Display for standard items
                col1, col2 = st.columns([3, 1])
                with col1:
                    # Create two columns within col1 - one for dropdown, one for smart filter
                    col1a, col1b = st.columns([3, 1])
                    
                    with col1a:
                        item_name = st.selectbox(
                            "Select Item", 
                            [''] + item_names, 
                            index=item_names.index(item['Item']) + 1 if item['Item'] in item_names else 0, 
                            key=f"edit_item_{idx}"
                        )
                        st.text(f"Item Description: {item_name}" if item_name else "")
                    
                    with col1b:
                        st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
                        if st.button("🔍 Change Item", key=f"smart_filter_{idx}"):
                            st.session_state["wizard_target_index"] = idx
                            st.switch_page("pages/wizard.py")

                            
                with col2:
                    quantity = st.text_input(
                        "Quantity", 
                        str(item['Quantity']), 
                        key=f"edit_qty_{idx}", 
                        placeholder="Input Quantity"
                    )
                    gst_applicable = st.checkbox(
                        "GST Applicable?", 
                        value=item.get('GST_Applicable', True), 
                        key=f"edit_standard_gst_{idx}"
                    )
                    
                    # Show unit rate below quantity
                    st.markdown(f"**Rate:** ₹{item['Unit Price']:.2f} per {item['Item Unit']}")
                    # Quantity Remarks section (for standard items)
            
                    # Check if a remark already exists
                    remark = item.get('Quantity_Remarks', '')
                    
                    # Change button label based on whether remark exists
                    button_label = "✏️ Edit Remark" if remark else "➕ Add Remark"                    
                    if st.button(button_label, key=f"add_qty_remark_{idx}"):
                        st.session_state.selected_items[idx]['show_remark_input'] = True
                    
                    # Show saved remark always (read-only view)
                    if remark and not item.get('show_remark_input', False):
                        st.info(f"📋 Quantity Remark: {remark}")
                    
                    # Show input box if editing
                    if item.get('show_remark_input', False):
                        new_remark = st.text_input("Edit Remark", value=remark, key=f"qty_remark_{idx}", max_chars=100)
                        if st.button("Save Remark", key=f"save_remark_{idx}"):
                            st.session_state.selected_items[idx]['Quantity_Remarks'] = break_long_words(new_remark)
                            st.session_state.selected_items[idx]['show_remark_input'] = False
                            st.session_state[f"expander_{idx}"] = False  # Collapse after saving remark
                            st.rerun()

                # Action buttons inside expander
                col1, col2, col3, col4, col5, col6 = st.columns([1, 1, 1, 1, 0.5, 1])
                with col1:
                    if st.button(f"🔁 Update", key=f"update_{idx}"):
                        if item_name and quantity:
                            try:
                                quantity = float(quantity)
                                if quantity > 0:
                                    item_data = data[data['Item Name'] == item_name].iloc[0]
                                    unit_price = item_data['Unit Price']
                                    unit = item_data['Item Unit']
                                    cost = round(quantity * unit_price, 2)
                                    st.session_state.selected_items[idx] = {
                                        'Item': item_name,
                                        'Quantity': quantity,
                                        'Unit Price': unit_price,
                                        'Item Unit': unit,
                                        'Cost': cost,
                                        'Type': 'Standard',
                                        'GST_Applicable': gst_applicable,
                                        'Quantity_Remarks': item.get('Quantity_Remarks', ''),  # Preserve existing remarks
                                        'show_remark_input': False  # Add this line
                                    }
                                    st.session_state[f"expander_{idx}"] = False  # Add this line to collapse
                                    st.rerun()
                            except ValueError:
                                st.error("Please enter a valid quantity")
                with col2:
                    if st.button(f"❌ Remove", key=f"remove_{idx}"):
                        remove_item(idx)
                with col3:
                    if st.button("⬆️ Move Up", key=f"move_up_sub_{idx}"):
                        move_item_up(idx)
                with col4:
                    if st.button("⬇️ Move Down", key=f"move_down_sub_{idx}"):
                        move_item_down(idx)
                with col5:
                    new_pos = st.text_input(
                        "", 
                        key=f"move_input_{idx}", 
                        placeholder="Move To", 
                        label_visibility="collapsed"
                    )
                with col6:    
                    if st.button("Move", key=f"move_button_{idx}"):
                        try:
                            target_index = int(new_pos) - 1
                            if target_index == idx:
                                st.info("Item already at that position.")
                            elif 0 <= target_index < len(st.session_state.selected_items):
                                move_item_to_position(idx, target_index)
                            else:
                                st.warning("Invalid position number.")
                        except ValueError:
                            st.error("Please enter a valid number.")        
    
    if st.button("☁️ Save Progress", use_container_width=True):
            save_progress_to_firebase(st.session_state.logged_in_username)
    # Add New Item or Subheading buttons
    button_col1, button_col2, button_col3, button_col4, button_col5, button_col6, button_col7,button_col8  = st.columns([2, 2, 2, 2, 2, 2, 2, 2])
    with button_col7:
        if st.button("🔽 Add from Dropdown", key="add_item_btn"):
            # Toggle add item section and hide others
            st.session_state.show_add_item = not st.session_state.get('show_add_item', False)
            st.session_state.show_wizard = False
            st.session_state.adding_subheading = False
            st.session_state.show_add_other = False
            st.session_state.show_templates = False  # Add this line
            st.session_state.show_upload = False    # Add this line
            st.session_state.show_local_templates = False
            st.rerun()
    with button_col1:
        if st.button("🔍 Smart Filter View", key="open_wizard"):
            # Toggle wizard and hide others
            st.session_state.show_wizard = not st.session_state.get('show_wizard', False)
            st.session_state.show_add_item = False
            st.session_state.adding_subheading = False
            st.session_state.show_add_other = False
            st.session_state.show_templates = False  # Add this line
            st.session_state.show_upload = False    # Add this line
            st.session_state.show_local_templates = False
            st.rerun()
    with button_col6:
        if st.button("📌 Add Subheading", key="add_subheading_btn"):
            # Toggle subheading and hide others
            st.session_state.adding_subheading = not st.session_state.get('adding_subheading', False)
            st.session_state.show_add_item = False
            st.session_state.show_wizard = False
            st.session_state.show_add_other = False
            st.session_state.show_templates = False  # Add this line
            st.session_state.show_upload = False    # Add this line
            st.session_state.show_local_templates = False
            st.rerun()
    with button_col5:
        if st.button("🧩 Custom Items", key="add_other_btn", type="secondary", 
                    help="Add custom items not in database", use_container_width=True):
            # Toggle other items section and hide others
            st.session_state.show_add_other = not st.session_state.get('show_add_other', False)
            st.session_state.show_add_item = False
            st.session_state.show_wizard = False
            st.session_state.adding_subheading = False
            st.session_state.show_templates = False  # Add this line
            st.session_state.show_upload = False    # Add this line
            st.session_state.show_local_templates = False
            st.rerun()
    with button_col2:
        if st.button("🌐 Global Templates", key="show_templates_btn"):
            st.session_state.show_templates = not st.session_state.get('show_templates', False)
            st.session_state.show_add_item = False
            st.session_state.show_wizard = False
            st.session_state.adding_subheading = False
            st.session_state.show_add_other = False
            st.session_state.show_upload = False    # Add this line
            st.session_state.show_local_templates = False
            st.rerun()
    with button_col4:
        # Replace your existing upload button with this:
        if st.button("📤 Upload Excel File", key="show_upload_btn"):
            st.session_state.show_upload = not st.session_state.get('show_upload', False)
            st.session_state.show_templates = False
            st.session_state.show_add_item = False
            st.session_state.show_wizard = False
            st.session_state.adding_subheading = False
            st.session_state.show_add_other = False
            st.session_state.show_local_templates = False
            st.rerun()
    with button_col3:
        if st.button("🏠 Local Templates", key="show_local_templates_btn"):
            st.session_state.show_local_templates = not st.session_state.get('show_local_templates', False)
            st.session_state.show_add_item = False
            st.session_state.show_wizard = False
            st.session_state.adding_subheading = False
            st.session_state.show_add_other = False
            st.session_state.show_upload = False
            st.session_state.show_templates = False
            st.rerun()
    with button_col8:
            if st.button("🔁 Update All Items", key="update_all2", 
                        help="Update all items with current values"):
                updated_count = update_all_items()
                st.rerun()        
    # Show Add Item section if toggled on
    if st.session_state.get('show_add_item', False):
        idx = len([i for i in st.session_state.selected_items if i.get("Type") != "Subheading"])
        with st.container():
            st.markdown(f"<div class='estimate-item'>", unsafe_allow_html=True)
            col1, col2 = st.columns([3, 1])
            with col1:
                item_name = st.selectbox(
                    "Select Item", 
                    [''] + item_names, 
                    key=f"new_item_{idx}"
                )
                st.text(f"Item Description: {item_name}" if item_name else "")
            with col2:
                quantity = st.text_input(
                    "Quantity", 
                    "1", 
                    key=f"new_qty_{idx}", 
                    placeholder="Input Quantity"
                )
                gst_applicable = st.checkbox(
                    "GST Applicable?", 
                    value=True, 
                    key=f"new_item_gst_{idx}"
                )
                if item_name != '':
                    item_data = data[data['Item Name'] == item_name].iloc[0]
                    unit_price = item_data['Unit Price']
                    unit = item_data['Item Unit']
                    st.text(f"Rate: {unit_price:.2f}/{unit}")
                    if quantity:
                        try:
                            qty = float(quantity)
                            if qty > 0:
                                total = qty * unit_price
                                st.text(f"Amount: {total:.2f}")
                        except ValueError:
                            st.text("Invalid quantity")

            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button(f"Add to Estimate", key=f"add_{idx}"):
                    if item_name and quantity:
                        try:
                            quantity = float(quantity)
                            if quantity > 0:
                                item_data = data[data['Item Name'] == item_name].iloc[0]
                                unit_price = item_data['Unit Price']
                                unit = item_data['Item Unit']
                                cost = round(quantity * unit_price, 2)
                                st.session_state.selected_items.append({
                                    'Item': item_name,
                                    'Quantity': quantity,
                                    'Unit Price': unit_price,
                                    'Item Unit': unit,
                                    'Cost': cost,
                                    'Type': 'Standard',
                                    'GST_Applicable': gst_applicable,
                                    'Quantity_Remarks': ""
                                })
                                st.session_state.show_add_item = False
                                st.rerun()
                        except ValueError:
                            st.error("Please enter a valid quantity")
            with col2:
                if st.button("✕ Cancel", key=f"cancel_add_{idx}", type="primary", 
                            help="Close without adding item", use_container_width=True):
                    st.session_state.show_add_item = False
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

    # Show Smart Filter if toggled on
    if st.session_state.get('show_wizard', False):
        show_item_wizard(wizard_data, handle_item_selection, st.session_state.selected_items)
        if st.button("✕ Close Wizard", key="close_wizard", type="primary"):
            st.session_state.show_wizard = False
            if 'show_wizard_for_edit' in st.session_state:
                st.session_state.show_wizard_for_edit = None
            st.rerun()
        # Show Templates section if toggled on
    # In the load_templates section, replace with this:
    # Then in the template loading section where items are added:
    if st.session_state.get('show_templates', False):
        template_data = load_templates()
        template_names = list(template_data.keys())
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        st.markdown("### 🌐 Global Templates")
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        num_columns = 3
        for i in range(0, len(template_names), num_columns):
            cols = st.columns(num_columns)
            for j in range(num_columns):
                if i + j < len(template_names):
                    template_name = template_names[i + j]
                    with cols[j]:
                        if st.button(f"{template_name}", key=f"template_btn_{template_name}"):
                            from openpyxl import load_workbook
                            template_path = f"Data Base/Global Templates/{template_name}.xlsx"
                            wb = load_workbook(template_path)
                            ws = wb.active
    
                            current_subheading = None
                            added_count = 0
    
                            for row_idx in range(1, ws.max_row + 1):
                                # Detect subheadings from merged cells
                                is_subheading = False
                                for merge in ws.merged_cells.ranges:
                                    if merge.min_row == row_idx and merge.max_row == row_idx:
                                        if merge.min_col <= 2 and merge.max_col >= 2:
                                            current_subheading = ws.cell(row=row_idx, column=merge.min_col).value
                                            is_subheading = True
                                            break
    
                                if is_subheading:
                                    st.session_state.selected_items.append({
                                        'Item': current_subheading,
                                        'Type': 'Subheading'
                                    })
                                    continue
    
                                item_name = ws.cell(row=row_idx, column=1).value
                                quantity_cell = ws.cell(row=row_idx, column=2).value
    
                                if not item_name:
                                    continue
    
                                # Extract quantity and remarks
                                remarks = ""
                                quantity = None
                                if quantity_cell is not None:
                                    quantity_str = str(quantity_cell).strip()
                                    if "(" in quantity_str and ")" in quantity_str:
                                        parts = quantity_str.split("(", 1)
                                        remarks = parts[1].split(")", 1)[0].strip()
                                        try:
                                            quantity = float(parts[0].strip())
                                        except ValueError:
                                            quantity = None
                                    else:
                                        try:
                                            quantity = float(quantity_str.split("(")[0].strip()) if quantity_str.split("(")[0].strip() else None
                                        except ValueError:
                                            quantity = None
    
                                main_item = data[data['Item Name'] == item_name]
    
                                if not main_item.empty and quantity is not None:
                                    # Standard item
                                    main_item = main_item.iloc[0]
                                    st.session_state.selected_items.append({
                                        'Item': item_name,
                                        'Quantity': quantity,
                                        'Unit Price': main_item['Unit Price'],
                                        'Item Unit': main_item['Item Unit'],
                                        'Cost': quantity * main_item['Unit Price'],
                                        'Type': 'Standard',
                                        'GST_Applicable': True,
                                        'Quantity_Remarks': remarks,
                                        'Subheading': current_subheading
                                    })
                                    added_count += 1
                                else:
                                    # Other item - only reflect name, leave other fields blank/zero
                                    st.session_state.selected_items.append({
                                        'Item': item_name,
                                        'Quantity': 0,
                                        'Unit Price': 0,
                                        'Item Unit': "",
                                        'Cost': 0,
                                        'Type': 'Other',
                                        'GST_Applicable': True,
                                        'Quantity_Remarks': remarks,
                                        'Subheading': current_subheading
                                    })
                                    added_count += 1
    
                            st.success(f"✅ Added {added_count} items from '{template_name}' template!")
                            st.session_state.show_templates = False
                            st.rerun()
    
        if st.button("✕ Cancel", key="cancel_template", type="primary"):
            st.session_state.show_templates = False
            st.rerun()

    # Show Local Templates section if toggled on
    # Then in the local template loading section where items are added:
    if st.session_state.get('show_local_templates', False):
        template_data = load_local_templates(username)
        template_names = list(template_data.keys())
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        st.markdown("### 🏠 Your Local Templates")
        num_columns = 3
        for i in range(0, len(template_names), num_columns):
            cols = st.columns(num_columns)
            for j in range(num_columns):
                if i + j < len(template_names):
                    template_name = template_names[i + j]
                    with cols[j]:
                        if st.button(f"📝 {template_name}", key=f"local_template_btn_{template_name}"):
                            from openpyxl import load_workbook
                            template_path = os.path.join("Data Base/Local Templates", username, f"{template_name}.xlsx")
                            wb = load_workbook(template_path)
                            ws = wb.active
    
                            current_subheading = None
                            added_count = 0
    
                            for row_idx in range(1, ws.max_row + 1):
                                # Detect subheadings from merged cells
                                is_subheading = False
                                for merge in ws.merged_cells.ranges:
                                    if merge.min_row == row_idx and merge.max_row == row_idx:
                                        if merge.min_col <= 2 and merge.max_col >= 2:
                                            current_subheading = ws.cell(row=row_idx, column=merge.min_col).value
                                            is_subheading = True
                                            break
    
                                if is_subheading:
                                    st.session_state.selected_items.append({
                                        'Item': current_subheading,
                                        'Type': 'Subheading'
                                    })
                                    continue
    
                                item_name = ws.cell(row=row_idx, column=1).value
                                quantity_cell = ws.cell(row=row_idx, column=2).value
    
                                if not item_name:
                                    continue
    
                                # Extract quantity and remarks
                                remarks = ""
                                quantity = None
                                if quantity_cell is not None:
                                    quantity_str = str(quantity_cell).strip()
                                    if "(" in quantity_str and ")" in quantity_str:
                                        parts = quantity_str.split("(", 1)
                                        remarks = parts[1].split(")", 1)[0].strip()
                                        try:
                                            quantity = float(parts[0].strip())
                                        except ValueError:
                                            quantity = None
                                    else:
                                        try:
                                            quantity = float(quantity_str.split("(")[0].strip()) if quantity_str.split("(")[0].strip() else None
                                        except ValueError:
                                            quantity = None
    
                                main_item = data[data['Item Name'] == item_name]
    
                                if not main_item.empty and quantity is not None:
                                    # Standard item
                                    main_item = main_item.iloc[0]
                                    st.session_state.selected_items.append({
                                        'Item': item_name,
                                        'Quantity': quantity,
                                        'Unit Price': main_item['Unit Price'],
                                        'Item Unit': main_item['Item Unit'],
                                        'Cost': quantity * main_item['Unit Price'],
                                        'Type': 'Standard',
                                        'GST_Applicable': True,
                                        'Quantity_Remarks': remarks,
                                        'Subheading': current_subheading
                                    })
                                    added_count += 1
                                else:
                                    # Other item - only reflect name, leave other fields blank/zero
                                    st.session_state.selected_items.append({
                                        'Item': item_name,
                                        'Quantity': 0,
                                        'Unit Price': 0,
                                        'Item Unit': "",
                                        'Cost': 0,
                                        'Type': 'Other',
                                        'GST_Applicable': True,
                                        'Quantity_Remarks': remarks,
                                        'Subheading': current_subheading
                                    })
                                    added_count += 1
    
                            st.success(f"✅ Added {added_count} items from '{template_name}' template!")
                            st.session_state.show_local_templates = False
                            st.rerun()
    
            if st.button("✕ Cancel", key="cancel_local_template", type="primary"):
                st.session_state.show_local_templates = False
                st.rerun()
    # Excel upload section
    if st.session_state.get('show_upload', False):
        try:
            sample_data = base64.b64encode(open("Side Bar Files/Sample.xlsx", "rb").read()).decode("utf-8")
        except FileNotFoundError:
            st.error("Sample file not found")
            sample_data = ""
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background-color:#f0f2f6; padding:12px; border-radius:8px; margin-bottom:16px;">
            <p style="margin:0 0 12px 0; font-size:14px; color:#333;">
            <b>Upload Guide:</b> You can upload either:<br>
            1. Estimates saved on Cloud, OR<br>
            2. Estimates generated by this app (auto-detected), OR<br>
            3. Custom Excel files with item names in first column and quantities in second column</p>
            <div style="display: flex; align-items: center; gap: 10px;">
                <span style="font-size:13px; color:#555;">Download sample format:</span>
                <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{sample_data}" 
                   download="Sample_Format.xlsx" style="text-decoration: none;">
                    <button style="background-color:#4CAF50; color:white; border:none; padding:6px 12px; 
                                border-radius:4px; font-size:13px; cursor:pointer;">
                        ⬇️ Download Sample Excel
                    </button>
                </a>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        with st.container():
            # 🔘 Toggle Firebase file load panel
            if st.button("☁️ Load From Cloud", key="trigger_load_firebase"):
                st.session_state["show_firebase_load"] = not st.session_state.get("show_firebase_load", False)
            
            
            
            # 🔄 Show Firebase file cards only when toggled on
            if st.session_state.get("show_firebase_load", False):
                username = st.session_state.get("logged_in_username")
                files = list_user_files(username)
            
                if not files:
                    st.warning("No saved estimates found.")
                else:
                    st.markdown("**📂 Saved Estimates**")
                    file_names = list(files.keys())
                    cards_per_row = 4
            
                    for i in range(0, len(file_names), cards_per_row):
                        cols = st.columns(cards_per_row)
                        for j in range(cards_per_row):
                            if i + j < len(file_names):
                                file_name = file_names[i + j]
                                with cols[j]:
                                    with st.container(border=True):
                                        st.markdown(f"🚰 {file_name}", unsafe_allow_html=True)
                                        btn_col1, btn_col2 = st.columns([2, 1])
                                        with btn_col1:
                                            if st.button("📥 Load", key=f"open_{file_name}"):
                                                excel_io = load_excel_from_firebase(username, file_name)
                                                if excel_io:
                                                    st.session_state["excel_uploader"] = excel_io
                                                    st.session_state["uploaded_file_name"] = file_name
                                                    st.session_state["show_firebase_load"] = False
                                                    st.rerun()
                                                else:
                                                    st.error("❌ Failed to load file from Firebase.")
                                        with btn_col2:
                                            if st.button("🗑️", key=f"delete_{file_name}"):
                                                db.reference(f'estimates/{username}/{file_name}').delete()
                                                st.success(f"❌")
                                                st.rerun()
                                                
            if st.session_state.get("show_firebase_load", False):
                if st.button("❌ Close", key="close_firebase_view"):
                    st.session_state["show_firebase_load"] = False
                    st.rerun()        
                            
        # 🔁 Check if a file is loaded from Firebase
        uploaded_file = st.session_state.get("excel_uploader")
        
        # If not loaded from Firebase, fallback to manual upload
        if not uploaded_file:
            # ✅ Use Firebase-loaded file if available
            uploaded_file = st.session_state.get("excel_uploader")
            
            # ✅ Otherwise show upload widget
            if not uploaded_file:
                uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"], key="manual_upload")
                if uploaded_file:
                    st.session_state["uploaded_file_name"] = uploaded_file.name

        
        if uploaded_file is not None:
            st.success(f"📄 Using: {st.session_state.get('uploaded_file_name', 'Loaded File')}")
            try:
                from openpyxl import load_workbook
                wb = load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                
                
                
                # Check if it's an estimate generated by this app
                is_app_estimate = False
                try:
                    if (ws['B3'].value == "Item Name" and ws['C3'].value == "Qty" and
                        ws['D3'].value == "Unit" and ws['E3'].value == "Rate" and
                        ws['F3'].value == "Total" and ws['G3'].value == "GST"):
                        is_app_estimate = True
                except:
                    pass
                
                items = []
                current_subheading = None
                
                if is_app_estimate:
                    # Preload fields into uploaded_excel_data (without touching input widgets)
                    estimate_note = ""
                    for row in range(ws.max_row, 0, -1):
                        cell_val = ws.cell(row=row, column=1).value
                        if cell_val and str(cell_val).strip():
                            estimate_note = str(cell_val)
                            break
                    
                    file_cell = ws["I1"].value
                    date_cell = ws["I2"].value
                    
                    file_name = str(file_cell).strip() if file_cell else ""
                    estimate_date = str(date_cell).strip() if date_cell else ""
                    
                    
                    st.session_state.uploaded_excel_data = {
                        'work_desc': str(ws['A1'].value) if ws['A1'].value else "",
                        'head_note': str(ws['A2'].value) if ws['A2'].value else "",
                        'estimate_note': estimate_note,
                        'file_name': file_name,
                        'estimate_date': estimate_date
                    }
                    
                    
                    # Initialize estimate_note as empty string
                    estimate_note = ""
                    
                    # Loop through all rows in column A to find the matching text
                    for row in range(1, ws.max_row):
                        cell_val = ws.cell(row=row, column=1).value
                        if cell_val and str(cell_val).strip() == "All Items should be as per ISI Standards":
                            next_row_val = ws.cell(row=row + 1, column=1).value
                            if next_row_val and str(next_row_val).strip():
                                estimate_note = str(next_row_val).strip()
                            break
                    
                    # Save to session state
                    st.session_state.uploaded_excel_data['estimate_note'] = estimate_note
                    st.session_state.estimate_note = estimate_note

                
                    # Find the "Subtotal" row by checking merged cells
                    subtotal_row = None
                    for merge in ws.merged_cells.ranges:
                        if merge.min_row == merge.max_row:  # Only consider row merges
                            first_cell = ws.cell(row=merge.min_row, column=merge.min_col)
                            if first_cell.value == "Subtotal":
                                subtotal_row = merge.min_row
                                break
                    
                    if subtotal_row is None:
                        st.error("Could not find 'Subtotal' row in the uploaded estimate")
                        return
                    
                    # Set the range to process (rows 3 to subtotal_row)
                    row_range = range(4, subtotal_row)
                else:
                    # For regular Excel files, process all rows (skip header if needed)
                    row_range = range(1, ws.max_row + 1)
                
                # Process all rows in the determined range
                for row in row_range:
                    # Check if this row has any horizontal merged cells (subheading)
                    is_subheading = False
                    for merge in ws.merged_cells.ranges:
                        if merge.min_row == row and merge.max_row == row:  # Horizontal merge only
                            if merge.min_col <= 2 and merge.max_col >= 2:  # Merge includes column B (or first column)
                                current_subheading = str(ws.cell(row=row, column=merge.min_col).value)
                                is_subheading = True
                                break
                    
                    if is_subheading:
                        items.append({
                            'Item Name': current_subheading,
                            'Type': 'Subheading',
                            'Merged': True
                        })
                        continue
                    
                    # Get cell values based on file type
                    if is_app_estimate:
                        item_name = str(ws.cell(row=row, column=2).value) if ws.cell(row=row, column=2).value else ""
                        quantity_cell = ws.cell(row=row, column=3).value
                        unit_cell = ws.cell(row=row, column=4).value
                        rate_cell = ws.cell(row=row, column=5).value
                        total_price_cell = ws.cell(row=row, column=6).value
                        gst_cell = ws.cell(row=row, column=7).value if ws.max_column >= 7 else "Yes"  # Default to Yes if column doesn't exist
                        
                        # Check GST applicability
                        gst_applicable = str(gst_cell).strip().lower() == "yes" if gst_cell else True  # Default to True if empty
                        
                        # Detect "Other" items in app-generated Excel
                        is_other_item = False
                        if unit_cell == "-" and rate_cell == "-":
                            is_other_item = True
                        elif not data[data['Item Name'] == item_name].empty:
                            is_other_item = False
                        else:
                            is_other_item = True
    
                        if is_other_item:
                            try:
                                remarks = ""
                                qty = 0.0
                                
                                # Fetch quantity as number and remarks from column H
                                qty = 0.0
                                remarks_cell = ws.cell(row=row, column=8)  # Column H
                                remarks = str(remarks_cell.value).strip() if remarks_cell and remarks_cell.value else ""
                                
                                if quantity_cell is not None:
                                    quantity_str = str(quantity_cell).strip()
                                    try:
                                        qty = float(quantity_str.split("(")[0].strip())
                                    except ValueError:
                                        qty = 0.0
                                
                                
                                unit = str(unit_cell).strip() if unit_cell and str(unit_cell).strip() not in ["", "-"] else ""
                                rate = float(str(rate_cell).strip()) if rate_cell and str(rate_cell).strip() not in ["", "-"] else 0.0
                                total = float(total_price_cell) if total_price_cell else qty * rate
                        
                                items.append({
                                    'Item Name': item_name,
                                    'Quantity': qty,
                                    'Unit': unit,
                                    'Rate': rate,
                                    'Total Price': total,
                                    'Type': 'Other',
                                    'Remarks': remarks,
                                    'Subheading': current_subheading,
                                    'GST_Applicable': gst_applicable
                                })
                        
                            except (ValueError, TypeError):
                                items.append({
                                    'Item Name': item_name,
                                    'Quantity': 0,
                                    'Unit': "",
                                    'Rate': 0,
                                    'Total Price': 0,
                                    'Type': 'Other',
                                    'Remarks': "",
                                    'Subheading': current_subheading,
                                    'GST_Applicable': True
                                })
                        else:
                            # Standard item processing
                            remarks_cell = ws.cell(row=row, column=8)  # Column H
                            remarks = str(remarks_cell.value).strip() if remarks_cell and remarks_cell.value else ""
                            
                            quantity = None
                            if quantity_cell is not None:
                                quantity_str = str(quantity_cell).strip()
                                try:
                                    quantity = float(quantity_str.split("(")[0].strip())
                                except ValueError:
                                    quantity = None
                            
                        
                            items.append({
                                'Item Name': item_name,
                                'Quantity': quantity,
                                'Unit': unit_cell,
                                'Rate': rate_cell,
                                'Total Price': total_price_cell,
                                'Type': 'Standard',
                                'Remarks': remarks,
                                'Subheading': current_subheading,
                                'GST_Applicable': gst_applicable
                            })
                    else:
                        # Original processing for non-app Excel files
                        item_name = ws.cell(row=row, column=1).value
                        quantity_cell = ws.cell(row=row, column=2).value
                        total_price_cell = ws.cell(row=row, column=3).value if ws.max_column >= 3 else None
                        
                        if not item_name:
                            continue
                        
                        # Extract quantity and remarks
                        remarks = ""
                        quantity = None
                        if quantity_cell is not None:
                            quantity_str = str(quantity_cell).strip()
                            if "(" in quantity_str and ")" in quantity_str:
                                parts = quantity_str.split("(", 1)
                                remarks = parts[1].split(")", 1)[0].strip()
                                try:
                                    quantity = float(parts[0].strip())
                                except ValueError:
                                    quantity = None
                            else:
                                try:
                                    quantity = float(quantity_str.split("(")[0].strip()) if quantity_str.split("(")[0].strip() else None
                                except ValueError:
                                    quantity = None
                        
                        total_price = 0.0
                        if total_price_cell is not None:
                            try:
                                total_price = float(total_price_cell)
                            except (ValueError, TypeError):
                                pass
                        
                        items.append({
                            'Item Name': item_name, 
                            'Quantity': quantity,
                            'Remarks': remarks,
                            'Total Price': total_price,
                            'Subheading': current_subheading,
                            'GST_Applicable': True  # Default to True for non-app estimates
                        })
                
                items_df = pd.DataFrame(items)
                
                if items_df.empty:
                    st.error("No valid items found in the uploaded file")
                    return
                
                # Display preview
                st.markdown("**Preview of uploaded items:**")
                st.dataframe(items_df.head())
                
                col1, col2 = st.columns([1, 1])
                with col1:
                    if st.button("Add Uploaded Items", key="add_uploaded_items"):
                        main_items_data = data
                        added_count = 0
                        
                        for _, row in items_df.iterrows():
                            if row.get('Type') == 'Subheading':
                                st.session_state.selected_items.append({
                                    'Item': row['Item Name'],
                                    'Type': 'Subheading'
                                })
                                continue
                                
                            item_name = row['Item Name']
                            quantity = row['Quantity']
                            remarks = row.get('Remarks', '')
                            gst_applicable = row.get('GST_Applicable', True)  # Default to True if not specified
                            
                            if row.get('Type') == 'Other':
                                # Add "Other" item with all extracted values
                                st.session_state.selected_items.append({
                                    'Item': item_name,
                                    'Quantity': float(row.get('Quantity', 0)),
                                    'Unit Price': float(row.get('Rate', 0)),
                                    'Item Unit': str(row.get('Unit', '')),
                                    'Cost': float(row.get('Total Price', 0)),
                                    'Type': 'Other',
                                    'GST_Applicable': gst_applicable,
                                    'Quantity_Remarks': remarks
                                })
                                added_count += 1
                            else:
                                # Standard item processing
                                main_item = main_items_data[main_items_data['Item Name'] == item_name]
                                
                                if not main_item.empty and quantity is not None:
                                    main_item = main_item.iloc[0]
                                    st.session_state.selected_items.append({
                                        'Item': item_name,
                                        'Quantity': float(quantity),
                                        'Unit Price': main_item['Unit Price'],
                                        'Item Unit': main_item['Item Unit'],
                                        'Cost': float(quantity) * main_item['Unit Price'],
                                        'Type': 'Standard',
                                        'GST_Applicable': gst_applicable,
                                        'Quantity_Remarks': remarks
                                    })
                                    added_count += 1
                                else:
                                    # Fallback to "Other" if item not found
                                    st.session_state.selected_items.append({
                                        'Item': item_name,
                                        'Quantity': float(quantity) if quantity is not None else 0,
                                        'Unit Price': 0,
                                        'Item Unit': "",
                                        'Cost': float(row.get('Total Price', 0)),
                                        'Type': 'Other',
                                        'GST_Applicable': gst_applicable,
                                        'Quantity_Remarks': remarks
                                    })
                                    added_count += 1
                        
                        st.success(f"Added {added_count} items from uploaded file!")
                        st.session_state.show_upload = False
                        st.session_state["excel_uploader"] = None
                        st.session_state["uploaded_file_name"] = None

                        st.rerun()
                
                with col2:
                    if st.button("✕ Cancel", key="cancel_upload", type="primary"):
                        st.session_state.show_upload = False
                        st.session_state["excel_uploader"] = None
                        st.session_state["uploaded_file_name"] = None
                        st.rerun()
                        
            except Exception as e:
                st.error(f"Error reading Excel file: {str(e)}")
        else:
            if st.button("✕ Cancel", key="cancel_upload_no_file", type="primary"):
                st.session_state.show_upload = False
                st.rerun()  
    # Show Subheading section if toggled on
    if st.session_state.get('adding_subheading', False):
        subheading = st.text_area("Enter Subheading", key="new_subheading")
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("Add Subheading to Estimate", key="confirm_subheading"):
                if subheading.strip():
                    st.session_state.selected_items.append({
                        'Item': subheading.strip(),
                        'Type': 'Subheading'
                    })
                    st.session_state.adding_subheading = False
                    st.success(f"Subheading '{subheading.strip()}' added!")
                    st.rerun()
                else:
                    st.warning("Please enter a valid subheading.")
        with col2:
            if st.button("✕ Cancel", key="cancel_subheading", type="primary"):
                st.session_state.adding_subheading = False
                st.rerun()

    # Show Add Other section if toggled on
    if st.session_state.get('show_add_other', False):
        with st.container():
            st.markdown(f"<div class='estimate-item'>", unsafe_allow_html=True)
            col1, col2, col3 = st.columns([5, 1, 1])
            with col1:
                item_name = st.text_area(
                    "Item Description", 
                    key=f"other_item_name",
                    placeholder="Enter Item Name",
                    label_visibility="collapsed",
                    height=150  # Adjust height as needed
                )
            with col2:
                quantity = st.text_input(
                    "Quantity", 
                    key=f"other_item_qty",
                    placeholder="Quantity",
                    label_visibility="collapsed"
                )
                unit = st.text_input(
                    "Unit", 
                    key=f"other_item_unit",
                    placeholder="Unit",
                    label_visibility="collapsed"
                )
                unit_price = st.text_input(
                    "Unit Rate (₹)", 
                    key=f"other_item_rate",
                    placeholder="Unit Rate",
                    label_visibility="collapsed"
                )
            with col3:    
                # Calculate and display total
                try:
                    qty = float(quantity) if quantity else 0
                    rate = float(unit_price) if unit_price else 0
                    total = qty * rate
                    st.markdown(f"**Total: ₹{total:,.2f}**")
                except ValueError:
                    st.warning("Please enter valid numbers for quantity and rate")
                    
                gst_applicable = st.checkbox(
                    "GST Applicable?", 
                    value=True,
                    key=f"other_item_gst"
                )
    
            col1a, col2a = st.columns([1, 1])
            with col1a:
                if st.button(f"Add Custom Item", key=f"add_other_item", use_container_width=True):
                    if item_name and quantity and unit and unit_price:
                        try:
                            qty = float(quantity)
                            rate = float(unit_price)
                            if qty > 0 and rate >= 0:
                                st.session_state.selected_items.append({
                                    'Item': item_name,
                                    'Quantity': qty,
                                    'Unit Price': rate,
                                    'Item Unit': unit,
                                    'Cost': qty * rate,
                                    'Type': 'Other',
                                    'GST_Applicable': gst_applicable
                                })
                                st.session_state.show_add_other = False
                                st.rerun()
                        except ValueError:
                            st.error("Please enter valid numbers for quantity and rate")
            with col2a:
                if st.button("✕ Cancel", key=f"cancel_other_item", type="primary", help="Close without adding item", use_container_width=True):
                    st.session_state.show_add_other = False
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

    # Totals and file generation
    if any(i.get("Type") != "Subheading" for i in st.session_state.selected_items):
        # Calculate base values
        total_cost, gst, _, _ = calculate_totals(0)
        
        # Calculate maximum allowed unforeseen (2.5% of total + gst)
        max_unforeseen = (total_cost + gst) * 0.025
        default_unforeseen = max_unforeseen  # Default to max allowed
        
        # Decide the rounding label
        rounding_label = ""
        
        if st.session_state.edit_final_total:
            rounding_label = "Final Total"
        else:
            option = st.session_state.get("rounding_option", "1000")
            if option == "none":
                rounding_label = "Final Total (No Rounding)"
            elif option == "100":
                rounding_label = "Final Total (Rounded to 100)"
            elif option == "1000":
                rounding_label = "Final Total (Rounded to 1,000)"
            else:
                rounding_label = "Final Total"
                
        # Initialize or get current unforeseen amount
        if 'unforeseen_amount' not in st.session_state:
            st.session_state.unforeseen_amount = default_unforeseen
        st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)
        st.markdown("""
        <style>
        .thin-banner {
            background: #5d89a3;
            padding: 2px 0px !important;
            border-top-left-radius: 13px;
            border-top-right-radius: 13px;
            border-bottom-left-radius: 0px;
            border-bottom-right-radius: 0px;
            box-shadow: 0 1px 4px rgba(0, 0, 0, 0.05);
            margin: 0px 0 6px 0;
            text-align: center;
            height: 30px;
            display: flex;
            justify-content: center;
            align-items: center;
        }
        .thin-banner h4 {
            color: white;
            font-size: 20px;
            font-weight: 600;
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1;
        }
        </style>
        <div class="thin-banner">
            <h4>SUMMARY</h4>
        </div>
    """, unsafe_allow_html=True)
        st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
        st.write(f"Subtotal: ₹{total_cost:,.2f}")
        st.write(f"GST (18% on taxable items): ₹{gst:,.2f}")
        
        # Create columns to control the input field width
        col1, col2, col3 = st.columns([2.5, 1.5, 18])  # Slightly adjusted column widths

        with col1:
            unforeseen_input = st.text_input(
                f"Unforeseen (max 2.5%): ₹{max_unforeseen:,.2f}",
                value=f"{st.session_state.unforeseen_amount:,.2f}",
                key="unforeseen_input"
            )
        
        with col2:
            st.markdown("<div style='margin-top: 1.75em'></div>", unsafe_allow_html=True)  # Push button down
            if st.button("🔁", key="reset_unforeseen"):
                st.session_state.unforeseen_amount = round(max_unforeseen, 2)
                st.session_state.manual_final_total = None  # Optional
                st.session_state.edit_final_total = False   # ⬅️ Force toggle OFF
                st.rerun()
        
        
        # Process and validate input
        try:
            entered_amount = float(unforeseen_input.replace(',', ''))
            
            if entered_amount > max_unforeseen:
                st.warning(f"Amount reduced to maximum allowed (2.5% of total + GST = ₹{max_unforeseen:,.2f})")
                unforeseen_amount = max_unforeseen
            else:
                unforeseen_amount = entered_amount
                
            st.session_state.unforeseen_amount = unforeseen_amount
            
        except ValueError:
            st.error("Please enter a valid number")
            unforeseen_amount = st.session_state.unforeseen_amount
        
        # Calculate final totals with the validated amount
        total_cost, gst, unforeseen, final_total = calculate_totals(int(unforeseen_amount))
        
        # Add toggle for manual final total
        col1, col2, col3, col4, col5 = st.columns([5, 2, 2, 2, 2])
        with col1:
            st.markdown(f"**{rounding_label}: ₹{final_total:,.3f}**")
        with col2:
            round_1000 = st.checkbox(
                "Round to ₹1,000", 
                value=st.session_state.get('rounding_option', '1000') == '1000',
                key="round_1000_cb",
                on_change=lambda: set_rounding_option('1000')
            )
        
        with col3:
            round_100 = st.checkbox(
                "Round to ₹100", 
                value=st.session_state.get('rounding_option', '1000') == '100',
                key="round_100_cb",
                on_change=lambda: set_rounding_option('100')
            )
        
        with col4:
            no_rounding = st.checkbox(
                "No Rounding", 
                value=st.session_state.get('rounding_option', '1000') == 'none',
                key="no_rounding_cb",
                on_change=lambda: set_rounding_option('none')
            )
        
        with col5:
            manual_edit = st.checkbox(
                "Edit Manually", 
                value=st.session_state.get('edit_final_total', False),
                key="manual_edit_cb",
                on_change=lambda: set_rounding_option('manual')
            )
            
        
        # Handle manual final total input
        if st.session_state.edit_final_total:
            # Create columns to control layout
            col1, col2 = st.columns([1, 3])
            
            with col1:
                # Calculate minimum allowed total (subtotal + GST rounded up)
                min_total = math.ceil(total_cost + gst)
                
                # Create a text input styled as narrow (without +/- buttons)
                manual_total_str = st.text_input(
                    "Enter Final Total (₹)",
                    value=f"{final_total:,.0f}",
                    key="manual_final_total_input",
                    help=f"Minimum: ₹{min_total:,.0f}, Maximum: ₹{math.ceil(total_cost + gst + max_unforeseen):,.0f}"
                )
                
                try:
                    # Parse the input (remove commas and convert to float)
                    manual_total = float(manual_total_str.replace(',', ''))
                    
                    # Validate range
                    if manual_total < min_total:
                        st.warning(f"Total cannot be less than ₹{min_total:,.0f}")
                        manual_total = min_total
                    elif manual_total > (total_cost + gst + max_unforeseen):
                        st.warning(f"Total cannot exceed ₹{math.ceil((total_cost + gst + max_unforeseen) / 100) * 100:,.0f}")
                        manual_total = math.ceil((total_cost + gst + max_unforeseen) / 100) * 100
                    
                    if manual_total != final_total:
                        new_unforeseen = manual_total - (total_cost + gst)
                        st.session_state.unforeseen_amount = new_unforeseen
                        st.session_state.manual_final_total = manual_total
                        st.rerun()  # ⬅️ Force UI to reflect changes

                        
                except ValueError:
                    st.error("Please enter a valid number")
                    manual_total = final_total
        else:
            # Reset to calculated total when toggle is off
            if st.session_state.manual_final_total is not None:
                st.session_state.manual_final_total = None

        def add_spec_watermark(pdf):
            """Add Icons/spec1.png at top left of the current page"""
            spec_path = "Icons/spec1.png"
            if os.path.exists(spec_path):
                image_width = 10  # Adjust as needed
                x = 5
                y = 5  # 30 mm from top; adjust if needed
                pdf.image(spec_path, x=x, y=y, w=image_width)
                pdf.set_font('Arial', 'I', 8)
                pdf.cell(0, 5, f'{pdf.page_no()}', 0, 0, 'R')

        
        # File generation buttons
        col1, col2, col3, col4, col5, col6 = st.columns([1, 1.5, 1.2, 1, 1, 1])  # Added a 4th column for preview
        with col3:
            if st.button("📊 Generate Excel", key="generate_excel", use_container_width=True):
                update_all_items()
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Border, Side
                from num2words import num2words
        
                wb = Workbook()
                ws = wb.active
                ws.title = "Estimate"
                
                # Insert file name and date into I1 and I2
                file_name = st.session_state.get("file_name", "")
                estimate_date = st.session_state.get("estimate_date", None)

                ws["I1"] = f"{file_name}"
                ws["I2"] = f"{estimate_date}"
        
                # Define alignments
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
                # === Heading ===
                ws.merge_cells('A1:H1')
                ws['A1'] = estimate_heading
                ws['A1'].alignment = center_align
                ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
        
                # === Head Note ===
                head_note = st.session_state.get('head_note', '')
                ws.merge_cells('A2:H2')
                ws['A2'] = head_note
                ws['A2'].alignment = center_align
                ws['A2'].font = ws['A2'].font.copy(italic=True)
                ws.row_dimensions[2].height = 30 if head_note.strip() else 15
        
                # === Table Headers ===
                headers = ["Sl.No", "Item Name", "Qty", "Unit", "Rate", "Total", "GST", "Remarks"]
                ws.append(headers)
        
                # === Item Rows ===
                row_num = 4
                serial = 1
                for item in st.session_state.selected_items:
                    if item.get("Type") == "Subheading":
                        ws.merge_cells(f'A{row_num}:H{row_num}')
                        ws[f'A{row_num}'] = f" {item['Item']}"
                        ws[f'A{row_num}'].alignment = center_align
                        row_num += 1
                    else:
                        qty = round(float(item['Quantity']), 2)
                        rate = round(float(item['Unit Price']), 2)
                        total_formula = f"=C{row_num}*E{row_num}"
                        gst_text = "Yes" if item.get("GST_Applicable", False) else "No"
                        remarks = item.get("Quantity_Remarks", "")
        
                        ws.append([
                            serial,
                            item['Item'],
                            qty,
                            item['Item Unit'],
                            rate,
                            total_formula,
                            gst_text,
                            remarks
                        ])
                        serial += 1
                        row_num += 1
        
                # === Totals with Excel Formulas ===
                start_data_row = 4
                end_data_row = row_num - 1
        
                # Subtotal
                ws.merge_cells(f'A{row_num}:E{row_num}')
                ws[f'A{row_num}'] = "Subtotal"
                ws[f'F{row_num}'] = f"=SUM(F{start_data_row}:F{end_data_row})"
                subtotal_row = row_num
                row_num += 1
        
                # GST
                ws.merge_cells(f'A{row_num}:E{row_num}')
                ws[f'A{row_num}'] = "GST (18%)"
                ws[f'F{row_num}'] = f"=F{subtotal_row}*0.18"
                gst_row = row_num
                row_num += 1
        
                # Unforeseen (use logic if needed)
                false_unforeseen = final_total - (total_cost + gst)
                ws.merge_cells(f'A{row_num}:E{row_num}')
                ws[f'A{row_num}'] = "Unforeseen"
                ws[f'F{row_num}'] = f"{false_unforeseen}"
                row_num += 1
        
                # Final Total
                ws.merge_cells(f'A{row_num}:E{row_num}')
                ws[f'A{row_num}'] = rounding_label
                ws[f'F{row_num}'] = final_total
                row_num += 1
        
                # === Amount in Words ===
                if final_total > 0:
                    amount_words = num2words(int(round(float(final_total))), lang='en_IN').title() + " Rupees Only"
                    ws.merge_cells(f'A{row_num}:H{row_num}')
                    ws[f'A{row_num}'] = f"Amount in Words: {amount_words}"
                    ws[f'A{row_num}'].alignment = center_align
                    ws[f'A{row_num}'].font = ws[f'A{row_num}'].font.copy(bold=True)
                    row_num += 1
        
                # === ISI Clause ===
                ws.merge_cells(f'A{row_num}:H{row_num}')
                ws[f'A{row_num}'] = "All Items should be as per ISI Standards"
                ws[f'A{row_num}'].alignment = center_align
                ws[f'A{row_num}'].font = ws[f'A{row_num}'].font.copy(italic=True)
                row_num += 1
        
                # === Estimate Note (if any) ===
                if hasattr(st.session_state, 'estimate_note') and st.session_state.estimate_note.strip():
                    ws.merge_cells(f'A{row_num}:H{row_num}')
                    ws[f'A{row_num}'] = st.session_state.estimate_note
                    ws[f'A{row_num}'].alignment = center_align
                    note_lines = len(st.session_state.estimate_note.split('\n')) + 1
                    ws.row_dimensions[row_num].height = note_lines * 15
                    row_num += 1
        
                # === Styling Borders ===
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
                for row in ws.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=8):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = center_align
        
                # Left align Item Names (Column B)
                for row in ws.iter_rows(min_row=4, max_row=row_num - 1):
                    row[1].alignment = left_align  # B column
        
                # === Set Column Widths ===
                ws.column_dimensions['A'].width = 8    # Sl.No
                ws.column_dimensions['B'].width = 60   # Item Name
                ws.column_dimensions['C'].width = 12   # Qty
                ws.column_dimensions['D'].width = 8    # Unit
                ws.column_dimensions['E'].width = 12   # Rate
                ws.column_dimensions['F'].width = 15   # Total
                ws.column_dimensions['G'].width = 8    # GST
                ws.column_dimensions['H'].width = 25   # Remarks
        
                # === Download Excel ===
                file_name_input = st.session_state.get("file_name", "").strip()
                work_desc_input = st.session_state.get("work_desc", "").strip()
                
                # Combine file name + work description
                # Remove newlines and other problematic characters
                import re
                sanitized_work_desc = re.sub(r'[\n\r\t]+', ' ', work_desc_input)
                sanitized_work_desc = re.sub(r'[\n\r\\/*?:"<>|]', '-', work_desc_input).strip()
                combined_name = f"{file_name_input} {sanitized_work_desc}".strip()
                safe_filename = re.sub(r'[\\/*?:"<>|]', '-', combined_name)
                excel_file = f"{safe_filename or 'Estimate'}.xlsx"
                
                # Sanitize combined name
                safe_filename = re.sub(r'[\\/*?:\"<>|]', '-', combined_name)
                excel_file = f"{safe_filename or 'Estimate'}.xlsx"
                wb.save(excel_file)
        
                with open(excel_file, "rb") as f:
                    st.download_button(
                        "⬇️ Download Excel",
                        f,
                        file_name=excel_file,
                        mime="application/vnd.ms-excel",
                        key="download_excel"
                    )

        with col1:
            if st.button("💾 Save to Cloud", use_container_width=True):
                try:
                    # 🔧 Define the Excel generation logic inline
                    from openpyxl import Workbook
                    from openpyxl.styles import Alignment, Border, Side
                    from num2words import num2words
                    import re
            
                    update_all_items()
            
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Estimate"
            
                    # Extract values
                    file_name = st.session_state.get("file_name", "")
                    estimate_date = st.session_state.get("estimate_date", "")
                    head_note = st.session_state.get('head_note', '')
                    estimate_note = st.session_state.get('estimate_note', '')
                    final_total = st.session_state.get('final_total', 0)
                    total_cost = st.session_state.get('total_cost', 0)
                    gst = st.session_state.get('gst', 0)
                    rounding_label = st.session_state.get('rounding_label', 'Final Total')
            
                    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
                    ws["I1"] = file_name
                    ws["I2"] = estimate_date
            
                    ws.merge_cells('A1:H1')
                    ws['A1'] = estimate_heading
                    ws['A1'].alignment = center_align
                    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
            
                    ws.merge_cells('A2:H2')
                    ws['A2'] = head_note
                    ws['A2'].alignment = center_align
                    ws['A2'].font = ws['A2'].font.copy(italic=True)
                    ws.row_dimensions[2].height = 30 if head_note.strip() else 15
            
                    headers = ["Sl.No", "Item Name", "Qty", "Unit", "Rate", "Total", "GST", "Remarks"]
                    ws.append(headers)
            
                    row_num = 4
                    serial = 1
                    for item in st.session_state.selected_items:
                        if item.get("Type") == "Subheading":
                            ws.merge_cells(f'A{row_num}:H{row_num}')
                            ws[f'A{row_num}'] = f" {item['Item']}"
                            ws[f'A{row_num}'].alignment = center_align
                            row_num += 1
                        else:
                            qty = round(float(item['Quantity']), 2)
                            rate = round(float(item['Unit Price']), 2)
                            total_formula = f"=C{row_num}*E{row_num}"
                            gst_text = "Yes" if item.get("GST_Applicable", False) else "No"
                            remarks = item.get("Quantity_Remarks", "")
            
                            ws.append([
                                serial,
                                item['Item'],
                                qty,
                                item['Item Unit'],
                                rate,
                                total_formula,
                                gst_text,
                                remarks
                            ])
                            serial += 1
                            row_num += 1
            
                    start_data_row = 4
                    end_data_row = row_num - 1
            
                    ws.merge_cells(f'A{row_num}:E{row_num}')
                    ws[f'A{row_num}'] = "Subtotal"
                    ws[f'F{row_num}'] = f"=SUM(F{start_data_row}:F{end_data_row})"
                    subtotal_row = row_num
                    row_num += 1
            
                    ws.merge_cells(f'A{row_num}:E{row_num}')
                    ws[f'A{row_num}'] = "GST (18%)"
                    ws[f'F{row_num}'] = f"=F{subtotal_row}*0.18"
                    gst_row = row_num
                    row_num += 1
            
                    unforeseen = final_total - (total_cost + gst)
                    ws.merge_cells(f'A{row_num}:E{row_num}')
                    ws[f'A{row_num}'] = "Unforeseen"
                    ws[f'F{row_num}'] = f"{unforeseen}"
                    row_num += 1
            
                    ws.merge_cells(f'A{row_num}:E{row_num}')
                    ws[f'A{row_num}'] = rounding_label
                    ws[f'F{row_num}'] = final_total
                    row_num += 1
            
                    if final_total > 0:
                        amount_words = num2words(int(round(float(final_total))), lang='en_IN').title() + " Rupees Only"
                        ws.merge_cells(f'A{row_num}:H{row_num}')
                        ws[f'A{row_num}'] = f"Amount in Words: {amount_words}"
                        ws[f'A{row_num}'].alignment = center_align
                        ws[f'A{row_num}'].font = ws[f'A{row_num}'].font.copy(bold=True)
                        row_num += 1
            
                    ws.merge_cells(f'A{row_num}:H{row_num}')
                    ws[f'A{row_num}'] = "All Items should be as per ISI Standards"
                    ws[f'A{row_num}'].alignment = center_align
                    ws[f'A{row_num}'].font = ws[f'A{row_num}'].font.copy(italic=True)
                    row_num += 1
            
                    if estimate_note.strip():
                        ws.merge_cells(f'A{row_num}:H{row_num}')
                        ws[f'A{row_num}'] = estimate_note
                        ws[f'A{row_num}'].alignment = center_align
                        note_lines = len(estimate_note.split('\n')) + 1
                        ws.row_dimensions[row_num].height = note_lines * 15
                        row_num += 1
            
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
                    for row in ws.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=8):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = center_align
            
                    for row in ws.iter_rows(min_row=4, max_row=row_num - 1):
                        row[1].alignment = left_align  # B column
            
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 60
                    ws.column_dimensions['C'].width = 12
                    ws.column_dimensions['D'].width = 8
                    ws.column_dimensions['E'].width = 12
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 8
                    ws.column_dimensions['H'].width = 25
            
                    # 🔽 Now save to Firebase
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
            
                    username = st.session_state.logged_in_username
                    import re

                    file_name_input = st.session_state.get("file_name", "").strip()
                    work_desc_input = st.session_state.get("work_desc", "").strip()
                    
                    # ✅ Sanitize work description
                    sanitized_work_desc = re.sub(r'[\n\r\t]+', ' ', work_desc_input)
                    sanitized_work_desc = re.sub(r'[.\n\r\\/*?:"<>|]', '-', work_desc_input).strip()
                    
                    # ✅ Combine and sanitize full filename (same as download logic)
                    combined_name = f"{file_name_input} {sanitized_work_desc}".strip()
                    safe_filename = re.sub(r'[.\\/*?:"<>|]', '-', combined_name)
                    filename_no_ext = safe_filename or "Estimate"  # Final fallback only if everything is blank
                    
                    # ✅ Save to Firebase using consistent filename
                    save_excel_to_firebase(username, filename_no_ext, output)
                    
                    st.success("✅ Saved")
            
                except Exception as e:
                    st.error(f"❌ Save failed: {e}")
        with col2:
            if st.button("📕 Generate PDF File", use_container_width=True):
                from num2words import num2words
                # Update all items silently
                update_all_items()
                from fpdf import FPDF
            
                pdf = FPDF()
                pdf.set_auto_page_break(auto=True, margin=15)
                
                def add_watermark(pdf):
                    """Function to add a diagonal watermark to every page"""
                    pdf.set_font("Arial", style='B', size=72)
                    pdf.set_text_color(180, 240, 230)  # Light magenta color for watermark
                
                    text = "GWD"
                    text_width = pdf.get_string_width(text)
                    text_height = 72  # Approximate height of the text
                
                    # Calculate the center of the page
                    page_width = pdf.w
                    page_height = pdf.h
                    center_x = page_width / 2
                    center_y = page_height / 2
                
                    # Rotate around the center of the page
                    pdf.rotate(45, x=center_x, y=center_y)
                
                    # Position text such that it is centered
                    x = center_x - (text_width / 2)
                    y = center_y + (text_height / 4)  # slight adjustment
                
                    pdf.text(x, y, text)
                
                    # Reset rotation to avoid affecting other content
                    pdf.rotate(0)
                    
                    pdf.set_text_color(0, 0, 0)  # Black color for the main content
                    
                
                # Replace the PDF head note section with this more robust version:

                pdf.set_auto_page_break(auto=True, margin=15)
                pdf.add_page()
                add_watermark(pdf)
                add_spec_watermark(pdf)
                
                # In the PDF generation section, replace the work description and head note handling with this:

                # Main content
                pdf.set_font("Arial", 'B', 16)
                pdf.set_text_color(0, 0, 0)
                
                # Add file info at top left (same formatting as user_info)
                pdf.set_font("Arial", '', 10)
                file_name = st.session_state.get("file_name", "")
                estimate_date = st.session_state.get("estimate_date", None)
                file_info = f"File No: {file_name}\nDate: {estimate_date}"
                pdf.set_xy(10, 15)  # Top left margin (same Y level as user_info)
                pdf.multi_cell(60, 5, file_info, 0, 'L')
                
                # Add user info at top right
                pdf.set_font("Arial", '', 10)
                user_info = f"User: {username}\nCost Index: {cost_index}"
                pdf.set_xy(pdf.w - 60, 15)  # Position at top right with some margin
                pdf.multi_cell(50, 5, user_info, 0, 'R')  # Right-aligned multi-cell for multiple lines
                
                # Start work description below user info (with sufficient space)
                pdf.set_y(40)  # Move down from top
                pdf.set_font("Arial", 'B', 16)
                
                # Process multi-line heading with proper line breaks
                heading_lines = estimate_heading.split('\n')
                heading_height = 0
                
                for line in heading_lines:
                    # Calculate width of heading text
                    heading_width = pdf.get_string_width(line)
                    
                    # If line is too wide for page (with 20mm margins on each side)
                    if heading_width > (pdf.w - 40):
                        # Split long lines into multiple lines
                        words = line.split()
                        current_line = ""
                        
                        for word in words:
                            test_line = f"{current_line} {word}" if current_line else word
                            if pdf.get_string_width(test_line) < (pdf.w - 40):
                                current_line = test_line
                            else:
                                # Draw the current line
                                pdf.cell(200, 10, txt=current_line, ln=True, align='C')
                                heading_height += 10
                                current_line = word
                        if current_line:
                            pdf.cell(200, 10, txt=current_line, ln=True, align='C')
                            heading_height += 10
                    else:
                        # Single line if it fits
                        pdf.cell(200, 10, txt=line, ln=True, align='C')
                        heading_height += 10
                
                # Calculate position for head note (after work description with some padding)
                head_note_y = pdf.get_y() + 10  # Add 10mm padding after work description
                
                # Add head note if it exists (only if we have space)
                if hasattr(st.session_state, 'head_note') and st.session_state.head_note.strip():
                    try:
                        # Define margins
                        left_margin = 20
                        right_margin = 20
                        table_width = pdf.w - left_margin - right_margin
                        
                        pdf.set_font("Arial", '', 10)
                        
                        # Position head note after work description
                        pdf.set_y(head_note_y)
                        
                        # Calculate width for head note (same as table width)
                        head_note_width = table_width
                        
                        # First calculate height needed
                        pdf.set_x(left_margin)
                        test_lines = pdf.multi_cell(
                            w=head_note_width,
                            h=5,
                            txt=st.session_state.head_note,
                            split_only=True
                        )
                        head_note_height = 5 * len(test_lines)
                        
                        # Check if we need a new page
                        if pdf.get_y() + head_note_height > pdf.h - 30:
                            pdf.add_page()
                            add_watermark(pdf)
                            add_spec_watermark(pdf)
                            pdf.set_y(40)  # Reset Y position after new page
                        
                        # Draw head note box with border
                        pdf.set_x(left_margin)
                        pdf.multi_cell(
                            w=head_note_width,
                            h=5,
                            txt=st.session_state.head_note,
                            border=1,
                            align='L'
                        )
                        
                        # Add space after head note
                        pdf.ln(10)
                        
                    except Exception as e:
                        st.error(f"Error adding head note to PDF: {str(e)}")
                
                # Define column widths
                col_widths = [10, 70, 20, 20, 20, 30]
                headers = ["Sl.No", "Item Name", "Qty", "Unit", "Rate", "Total"]
                
                # Calculate total table width and left margin for centering
                table_width = sum(col_widths)
                left_margin = (pdf.w - table_width) / 2
                
                def split_text(text, max_width):
                    """Split text into multiple lines based on available width"""
                    if not isinstance(text, str):
                        text = str(text)
                    lines = []
                    words = text.split()
                    current_line = ""
            
                    for word in words:
                        test_line = current_line + " " + word if current_line else word
                        if pdf.get_string_width(test_line) < max_width - 2:
                            current_line = test_line
                        else:
                            lines.append(current_line)
                            current_line = word
                    if current_line:
                        lines.append(current_line)
                    return lines
            
                def calculate_max_lines(row_data):
                    """Calculate maximum lines needed for any cell in the row"""
                    max_lines = 1
                    for i, text in enumerate(row_data):
                        lines = split_text(str(text), col_widths[i])
                        if len(lines) > max_lines:
                            max_lines = len(lines)
                    return max_lines
            
                def draw_table_header():
                    """Draw the table header on new pages"""
                    pdf.set_font("Arial", 'B', 10)
                    x_start = left_margin  # Use calculated left margin
                    y_start = pdf.get_y()
                    pdf.rect(x_start, y_start, table_width, 6)  # Header border
            
                    for i in range(1, len(col_widths)):
                        pdf.line(
                            x_start + sum(col_widths[:i]), y_start,
                            x_start + sum(col_widths[:i]), y_start + 6
                        )
            
                    for i, header in enumerate(headers):
                        pdf.set_xy(x_start + sum(col_widths[:i]), y_start)
                        pdf.cell(col_widths[i], 6, header, 0, 0, 'C')
            
                    pdf.set_y(y_start + 6)
            
                pdf.ln(10)
                draw_table_header()
                pdf.set_font("Arial", '', 10)
            
                serial = 1
                for item in st.session_state.selected_items:
                    # Check if we need a new page (with buffer for row height)
                    if pdf.get_y() + 20 > pdf.h - 30:  # Increased buffer to 20
                        pdf.add_page()
                        add_watermark(pdf)
                        add_spec_watermark(pdf)
                        draw_table_header()
                        pdf.set_font("Arial", '', 10)  # Reset font after header
            
                    if item.get("Type") == "Subheading":
                        # Calculate height needed (using fixed 6mm per line)
                        available_width = table_width - 10  # 5mm margin each side
                        subheading_lines = split_text(item['Item'], available_width)
                        subheading_height = 6 * len(subheading_lines)
                        
                        # Page break check
                        if pdf.get_y() + subheading_height > pdf.h - 30:
                            pdf.add_page()
                            add_watermark(pdf)
                            add_spec_watermark(pdf)
                            draw_table_header()
                        
                        # Draw border
                        x_start = left_margin
                        y_start = pdf.get_y()
                        pdf.rect(x_start, y_start, table_width, subheading_height)
                        
                        # Print text with margins
                        pdf.set_font("Arial", 'B', 10)
                        for i, line in enumerate(subheading_lines):
                            pdf.set_x(x_start + 5)  # 5mm left margin
                            pdf.cell(table_width - 10, 6, line, 0, 0, 'C')  # Centered in remaining width
                            if i < len(subheading_lines) - 1:
                                pdf.ln()
                        
                        pdf.set_y(y_start + subheading_height)
                        pdf.set_font("Arial", '', 10)
                        continue
                    
                    gst_applicable = item.get('GST_Applicable', True)
            
                    # In the PDF generation section, modify the item processing logic:

                    if item.get("Type") == "Other":
                        rate_text = f"{item['Unit Price']:.2f}" if 'Unit Price' in item else "-"
                        unit_text = item.get('Item Unit', '-')
                        remark = item.get('Quantity_Remarks', '')
                        if remark:
                            qty_text = f"{item.get('Quantity', '-')} ({remark})"
                        else:
                            qty_text = f"{item.get('Quantity', '-')}"
                    # Replace it with this:
                    else:
                        rate_text = f"{item['Unit Price']:.2f}"
                        unit_text = item['Item Unit']
                        remark = item.get('Quantity_Remarks', '')
                        # Format quantity to remove trailing zeros
                        qty_value = item['Quantity']
                        qty_str = f"{qty_value:.4f}".rstrip('0').rstrip('.') if '.' in f"{qty_value:.4f}" else f"{qty_value}"
                        if remark:
                            qty_text = f"{qty_str} ({remark})"
                        else:
                            qty_text = qty_str
                    
                    total_text = f"{item['Cost']:.2f}"
                    if not gst_applicable:
                        total_text += " (No GST)"
            
                    row_data = [
                        str(serial),
                        item['Item'],
                        qty_text,
                        unit_text,
                        rate_text,
                        total_text
                    ]
            
                    x_row_start = left_margin
                    y_row_start = pdf.get_y()
            
                    max_lines = calculate_max_lines(row_data)
                    row_height = 6 * max_lines
                    
                    # Ensure we have space for this row
                    if pdf.get_y() + row_height > pdf.h - 30:
                        pdf.add_page()
                        add_watermark(pdf)
                        add_spec_watermark(pdf)
                        draw_table_header()
                        pdf.set_font("Arial", '', 10)
                        x_row_start = left_margin
                        y_row_start = pdf.get_y()
            
                    # Draw the row border
                    pdf.rect(x_row_start, y_row_start, table_width, row_height)
                    
                    # Draw vertical lines
                    for i in range(1, len(col_widths)):
                        pdf.line(
                            x_row_start + sum(col_widths[:i]), y_row_start,
                            x_row_start + sum(col_widths[:i]), y_row_start + row_height
                        )
            
                    # For the first column (serial number) - keep centered
                    if item.get("Type") == "Other":
                        # Calculate circle position and size
                        r = 4  # Radius of the circle
                        x = x_row_start + col_widths[0]/2  # Center of the first column
                        y = y_row_start + row_height/2     # Vertical center of the row
                        
                        # Draw circle
                        pdf.ellipse(x - r, y - r, r * 2, r * 2)
                        
                        # Print serial number centered in the circle
                        pdf.set_xy(x_row_start, y_row_start)
                        pdf.cell(col_widths[0], row_height, str(serial), 0, 0, 'C')
                    else:
                        # Standard item - just print the serial number
                        pdf.set_xy(x_row_start, y_row_start)
                        pdf.cell(col_widths[0], row_height, str(serial), 0, 0, 'C')
                
                    # For the item name column (second column) - left-justified
                    pdf.set_xy(x_row_start + col_widths[0], y_row_start)
                    cell_lines = split_text(str(item['Item']), col_widths[1])
                    vertical_offset = (row_height - (6 * len(cell_lines))) / 2
                    
                    for line in cell_lines:
                        pdf.set_xy(x_row_start + col_widths[0], y_row_start + vertical_offset)
                        pdf.cell(col_widths[1], 6, line, 0, 0, 'L')  # Changed to 'L' for left alignment
                        vertical_offset += 6
                
                    # For remaining columns (keep centered)
                    for i, text in enumerate(row_data[2:], 2):  # Start from index 2 (rate)
                        pdf.set_xy(x_row_start + sum(col_widths[:i]), y_row_start)
                        cell_lines = split_text(str(text), col_widths[i])
                        vertical_offset = (row_height - (6 * len(cell_lines))) / 2
                        
                        for line in cell_lines:
                            pdf.set_xy(x_row_start + sum(col_widths[:i]), y_row_start + vertical_offset)
                            pdf.cell(col_widths[i], 6, line, 0, 0, 'C')  # Keep centered
                            vertical_offset += 6
            
                    pdf.set_y(y_row_start + row_height)
                    serial += 1
            
                # === STEP 1: Estimate dynamic content height ===

                note_height = 0
                if 'estimate_note' in st.session_state and st.session_state.estimate_note.strip():
                    pdf.set_font("Arial", '', 9)
                    note_lines = pdf.multi_cell(0, 5, st.session_state.estimate_note, split_only=True)
                    note_height = 5 * len(note_lines) + 2  # +2mm padding
                
                # Fixed content heights
                summary_height = 4 * 8  # 4 summary rows
                amount_words_height = 8
                isi_clause_height = 8
                signature_height = 35
                padding = 10  # buffer space before signature
                
                # Total content to print
                total_block_height = summary_height + amount_words_height + isi_clause_height + note_height + signature_height + padding
                
                # === STEP 2: Check space available ===
                available_space = pdf.h - pdf.get_y() - 20  # 20mm bottom margin
                if available_space < total_block_height:
                    pdf.add_page()
                    add_watermark(pdf)
                    add_spec_watermark(pdf)
                    pdf.ln(10)  # optional margin on new page

                # Summary Section
                false_unforeseen = final_total - (total_cost + gst)
            
                summary_data = [
                    ("Subtotal", f"{total_cost:.2f}"),
                    ("GST (18%)", f"{gst:.2f}"),
                    ("Unforeseen", f"{false_unforeseen:.2f}"),
                    (rounding_label, f"{final_total:.2f}")
                ]
            
                for label, value in summary_data:
                    row_height = 8
                    if pdf.get_y() + row_height > pdf.h - 30:
                        pdf.add_page()
                        add_watermark(pdf)
                        add_spec_watermark(pdf)
                        pdf.set_font("Arial", '', 10)  # Reset the correct font and size
                
                    x = left_margin
                    y = pdf.get_y()
                    
                    # Set bold font for summary items
                    pdf.set_font("Arial", 'B', 10)  # Changed to bold
            
                    # Draw both label and value in the same row
                    pdf.set_xy(x, y)
                    pdf.cell(sum(col_widths[:-1]), row_height, label, border=1, align='C')
                
                    pdf.set_xy(x + sum(col_widths[:-1]), y)
                    pdf.cell(col_widths[-1], row_height, value, border=1, align='C')
                
                    # Move to next line
                    pdf.set_y(y + row_height)
                
                 # Add amount in words (merged row)
                row_height = 8
                if pdf.get_y() + row_height > pdf.h - 30:
                    pdf.add_page()
                    add_watermark(pdf)
                    add_spec_watermark(pdf)
                
                # Set x position to left margin to center the cell
                pdf.set_x(left_margin)
                pdf.set_font("Arial", 'B', 10)
                
                # Calculate the amount in words
                amount_words = "Amount in Words: " + num2words(int(round(float(final_total))), lang='en_IN').title() + " Rupees Only"
                
                # First calculate how many lines we need
                test_lines = pdf.multi_cell(
                    w=table_width,
                    h=6,  # line height
                    txt=amount_words,
                    split_only=True
                )
                lines_needed = len(test_lines)
                required_height = 6 * lines_needed  # 6mm per line
                
                # Check if we need more space
                if pdf.get_y() + required_height > pdf.h - 30:
                    pdf.add_page()
                    add_watermark(pdf)
                    add_spec_watermark(pdf)
                    pdf.set_x(left_margin)
                
                # Now actually draw the cell with the correct height
                pdf.multi_cell(
                    w=table_width,
                    h=6,  # line height
                    txt=amount_words,
                    border=1,
                    ln=1,  # move to next line after
                    align='C'
                )
                
                # Add ISI standards note (merged row)
                pdf.set_x(left_margin)  # Set x position to left margin to center the cell
                pdf.set_font("Arial", 'I', 10)
                pdf.cell(table_width, row_height, "All Items should be as per ISI Standards", border=1, ln=1, align='C')
                pdf.set_font("Arial", '', 10)  # Reset font
                
                
                # In the PDF generation section, modify the note section:
                if 'estimate_note' in st.session_state and st.session_state.estimate_note.strip():
                    # Calculate height needed for note (approx 6mm per line)
                    pdf.set_font("Arial", '', 9)  # Slightly smaller font for longer notes
                    note_lines = pdf.multi_cell(0, 5, st.session_state.estimate_note, split_only=True)
                    note_height = 5 * len(note_lines)  # 5mm per line
                    
                    # Check if we need a new page for the note
                    if pdf.get_y() + note_height > pdf.h - 60:  # 60mm buffer for signatures
                        pdf.add_page()
                    
                    # Add minimal gap (2mm)
                    pdf.ln(2)
                    
                    # Draw note box with border
                    pdf.set_x(left_margin)
                    pdf.multi_cell(
                        table_width, 
                        5,  # Line height
                        st.session_state.estimate_note,
                        border=1,
                        align='C'
                    )
                    
                    # Reset font for remaining content
                    pdf.set_font("Arial", '', 10)
                    
                    
                # Signature Area (Immediately after last content)
                pdf.ln(st.session_state.signature_height)  # Use the configurable space
                
                # Capture current Y after all previous content (note or ISI clause)
                signature_y = pdf.get_y()
                
                # Check if we need a new page using the configurable values
                if signature_y + st.session_state.signature_block_height > pdf.h - st.session_state.bottom_margin:
                    pdf.add_page()
                    add_watermark(pdf)
                    add_spec_watermark(pdf)
                    signature_y = pdf.get_y()
                
                # Signature labels with equal spacing
                pdf.set_font("Arial", 'B', 10)
                
                # Calculate positions based on configured height
                signature_block_height = st.session_state.signature_block_height
                signature_text_height = 10  # Height for the text lines
                vertical_padding = (signature_block_height - (2 * signature_text_height)) / 2
                
                # Assistant Engineer (Left)
                pdf.set_xy(20, signature_y)
                pdf.cell(50, signature_text_height, "Assistant Engineer", ln=True, align='C')
                pdf.set_xy(20, signature_y + signature_text_height/2)
                pdf.cell(50, signature_text_height, "(Seal & Signature)", ln=True, align='C')
                
                # Assistant Executive Engineer (Center)
                center_x = pdf.w/2 - 25
                pdf.set_xy(center_x, signature_y)
                pdf.cell(50, signature_text_height, "Assistant Executive Engineer", ln=True, align='C')
                pdf.set_xy(center_x, signature_y + signature_text_height/2)
                pdf.cell(50, signature_text_height, "(Seal & Signature)", ln=True, align='C')
                
                # District Officer (Right)
                right_x = pdf.w - 70
                pdf.set_xy(right_x, signature_y)
                pdf.cell(50, signature_text_height, "District Officer", ln=True, align='C')
                pdf.set_xy(right_x, signature_y + signature_text_height/2)
                pdf.cell(50, signature_text_height, "(Seal & Signature)", ln=True, align='C')
                
                # Add horizontal line above signatures
                pdf.set_line_width(0.5)
                pdf.line(20, signature_y, pdf.w - 20, signature_y)
        

                
                # Save and offer download
                file_name_input = st.session_state.get("file_name", "").strip()
                work_desc_input = st.session_state.get("work_desc", "").strip()
                
                # Combine File Name + Work Description
                import re
                sanitized_work_desc = re.sub(r'[\n\r\t]+', ' ', work_desc_input)
                sanitized_work_desc = re.sub(r'[\n\r\\/*?:"<>|]', '-', work_desc_input).strip()
                combined_name = f"{file_name_input} {sanitized_work_desc}".strip()
                safe_filename = re.sub(r'[\\/*?:"<>|]', '-', combined_name)
                pdf_file = f"{safe_filename or 'Estimate'}.pdf"
                
                # Sanitize for filename use
                safe_filename = re.sub(r'[\\/*?:\"<>|]', '-', combined_name)
                
                # Create PDF file
                pdf_file = f"{safe_filename or 'Estimate'}.pdf"
                pdf.output(pdf_file)
            
                with open(pdf_file, "rb") as f:
                    st.download_button(
                        label="⬇️ Download PDF",
                        data=f,
                        file_name=pdf_file,
                        mime="application/pdf"
                    )
            # Toggle button to show/hide signature area settings
            show_signature_settings = st.toggle("Show PDF Signature Area Settings")
            
            if show_signature_settings:
                # Create three equal-width columns
                pdf_param_col1, pdf_param_col2, pdf_param_col3 = st.columns(3)
            
                # Initialize parameters if not in session state
                if 'signature_height' not in st.session_state:
                    st.session_state.signature_height = 30
                if 'signature_block_height' not in st.session_state:
                    st.session_state.signature_block_height = 35
                if 'bottom_margin' not in st.session_state:
                    st.session_state.bottom_margin = 20
            
                with pdf_param_col1:
                    st.session_state.signature_height = st.number_input(
                        " ",
                        min_value=5,
                        max_value=100,
                        value=st.session_state.signature_height,
                        step=5,
                        key="signature_height_input",
                        help="Vertical space before signatures (mm)"
                    )
            
                with pdf_param_col2:
                    st.session_state.signature_block_height = st.number_input(
                        " ",
                        min_value=5,
                        max_value=100,
                        value=st.session_state.signature_block_height,
                        step=5,
                        key="signature_block_height_input",
                        help="Height of signature block (mm)"
                    )
            
                with pdf_param_col3:
                    st.session_state.bottom_margin = st.number_input(
                        " ",
                        min_value=5,
                        max_value=50,
                        value=st.session_state.bottom_margin,
                        step=5,
                        key="bottom_margin_input",
                        help="Page bottom margin (mm)"
                    )
        with col4:
            if st.button("👁️ Preview", key="preview_estimate", use_container_width=True):
                st.session_state.show_preview = not st.session_state.get('show_preview', False)
                st.rerun()
        with col5:
            if st.button("🗑️ Clear All", key="clear_all", 
                        help="Remove all items and start fresh", use_container_width=True):
                st.session_state.selected_items = []
                st.session_state.item_count = 0
                st.session_state.adding_subheading = False
                st.session_state.show_wizard = False
                st.session_state.show_add_item = False
                st.session_state.show_add_other = False
                # Use st.query_params to clear the fields
                st.query_params.clear()
                st.query_params.update({
                    "clear_all": "true",
                    "work_desc": "",
                    "head_note": "",
                    "estimate_note": ""
                })
                st.rerun()
              
        with col6:
            if st.button("🔁 Update All", key="update_all1", 
                        help="Update all items with current values", use_container_width=True):
                updated_count = update_all_items()
                st.rerun()
            
        
    # Add this right after the totals section but before the "else" for "No items added"
    if st.session_state.get('show_preview', False) and any(i.get("Type") != "Subheading" for i in st.session_state.selected_items):
        st.markdown("""
        <style>
        .thin-banner {
            background: #5d89a3;
            padding: 2px 0px !important;
            border-top-left-radius: 13px;
            border-top-right-radius: 13px;
            border-bottom-left-radius: 0px;
            border-bottom-right-radius: 0px;
            box-shadow: 0 1px 4px rgba(0, 0, 0, 0.05);
            margin: 0px 0 6px 0;
            text-align: center;
            height: 30px;
            display: flex;
            justify-content: center;
            align-items: center;
        }
        .thin-banner h4 {
            color: white;
            font-size: 20px;
            font-weight: 600;
            margin: 0 !important;
            padding: 0 !important;
            line-height: 1;
        }
        </style>
        <div class="thin-banner">
            <h4>PREVIEW</h4>
        </div>
    """, unsafe_allow_html=True)
        st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)
        
        # Create a preview dataframe
        preview_data = []
        for idx, item in enumerate(st.session_state.selected_items):
            if item.get("Type") == "Subheading":
                preview_data.append({
                    "Item": f"📌 {item['Item']}",
                    "Quantity": "",
                    "Unit": "",
                    "Rate": "",
                    "Amount": ""
                })
            else:
                # In the preview section, replace the "Other" item handling with:
                if item.get("Type") == "Other":
                    remark = f" ({item['Quantity_Remarks']})" if item.get('Quantity_Remarks') else ""
                    preview_data.append({
                        "Item": f"🔹 {item['Item']}",
                        "Quantity": f"{item['Quantity']}{remark}",
                        "Unit": item['Item Unit'],
                        "Rate": f"₹{item['Unit Price']:,.2f}",
                        "Amount": f"₹{item['Cost']:,.2f}"
                    })
                
                else:
                    remark = f" ({item['Quantity_Remarks']})" if item.get('Quantity_Remarks') else ""
                    preview_data.append({
                        "Item": item['Item'],
                        "Quantity": f"{item['Quantity']}{remark}",
                        "Unit": item['Item Unit'],
                        "Rate": f"₹{item['Unit Price']:,.2f}",
                        "Amount": f"₹{item['Cost']:,.2f}"
                    })
        
        # Convert to dataframe and display
        preview_df = pd.DataFrame(preview_data)
        st.dataframe(
            preview_df,
            use_container_width=True,
            hide_index=True
        )
        
        # Add totals to the preview
        total_cost, gst, unforeseen, final_total = calculate_totals()
        false_unforeseen = final_total - (total_cost + gst)
        st.markdown(f"""
        **Subtotal:** ₹{total_cost:,.2f}  
        **GST (18%):** ₹{gst:,.2f}  
        **Unforeseen (max 2.5%):** ₹{false_unforeseen:,.2f}  
        **{rounding_label}: ₹{final_total:,.3f}**
        """)
        
        col1, col2= st.columns([1, 4])
        with col1:
            if st.button("Close Preview", key="close_preview"):
                st.session_state.show_preview = False
                st.rerun()
# Check authentication
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.logged_in_username = None

# Load credentials
try:
    credentials_df = load_credentials("Data Base/items.xltm")
except Exception as e:
    st.error(f"Error loading credentials: {str(e)}")
    st.stop()

if st.session_state.authenticated:
    main_app()
else:
    login_page(credentials_df)
    
# Replace the existing sidebar CSS with this:
st.sidebar.markdown("""
<style>
    /* Remove the arrow button styling */
    section[data-testid="stSidebar"] div[data-testid="stToolbar"] {
        display: none !important;
    }
    
    /* Main container styling */
    section[data-testid="stSidebar"] > div {
        padding-top: 0 !important;
    }
    
    /* Button styling - compact and professional */
    section[data-testid="stSidebar"] button {
        width: 100% !important;
        margin: 0px 0 !important;
        padding: 0px !important;
        font-size: 14px !important;
        border-radius: 4px !important;
        border: 1px solid #2387eb !important;
        background-color: #e8f2fc !important;
        color: #333 !important;
        transition: all 0.2s !important;
        box-shadow: none !important;
    }
    
    /* Download/secondary buttons - green style */
    section[data-testid="stSidebar"] .stDownloadButton button,
    section[data-testid="stSidebar"] .pump-selector-btn button {
        background-color: #4CAF50 !important;
        color: white !important;
        border: 1px solid #2E7D32 !important;
    }
    
    /* Hover states */
    section[data-testid="stSidebar"] button:hover:not(.stDownloadButton button, .pump-selector-btn button) {
        background-color: #d0e3f7 !important;
        border-color: #1a6fbb !important;
    }
    
    section[data-testid="stSidebar"] .stDownloadButton button:hover,
    section[data-testid="stSidebar"] .pump-selector-btn button:hover {
        background-color: #43a047 !important;
        border-color: #1b5e20 !important;
    }
    
    /* Active states */
    section[data-testid="stSidebar"] button:active:not(.stDownloadButton button, .pump-selector-btn button) {
        background-color: #b8d4f0 !important;
    }
    
    section[data-testid="stSidebar"] .stDownloadButton button:active,
    section[data-testid="stSidebar"] .pump-selector-btn button:active {
        background-color: #388e3c !important;
    }
    
    /* Dropdown styling */
    section[data-testid="stSidebar"] .stSelectbox select {
        padding: 0px 0px !important;
        font-size: 14px !important;
        margin: 0px 0 0px 0 !important;
    }
    
    /* Section headers */
    section[data-testid="stSidebar"] .st-expanderHeader {
        padding: 6px 8px !important;
        margin: 2px 0 !important;
    }
    
    /* Remove extra padding around buttons */
    section[data-testid="stSidebar"] > div > div > div > div {
        padding: 0 !important;
    }
    
    /* Pump selector link button */
    .pump-selector-btn button {
        margin-top: 0px !important;
        margin-bottom: 0px !important;
    }
</style>
""", unsafe_allow_html=True)


if st.session_state.get('authenticated', False):
    # Add the DSR download button and dropdowns
    # DSR/DAR button
    if st.sidebar.button("Download DSR/DAR"):
        toggle_section('show_dsr_options')
        st.rerun()  # Force immediate update

    if st.session_state.get('show_dsr_options', False):
        # Year selection
        selected_year = st.sidebar.selectbox("Select Year", ["2018", "2021"])
        
        # Document type selection
        doc_type = st.sidebar.selectbox("Select Document Type", ["DSR", "DAR"])
        
        # Volume selection
        volume = st.sidebar.selectbox("Select Volume", ["Vol 1", "Vol 2"])
        
        # Construct the file path
        file_path = f"Side Bar Files/DSR/{selected_year}/{doc_type}/{volume}.pdf"
        
        # Display download button
        try:
            with open(file_path, "rb") as file:
                st.sidebar.download_button(
                    label=f"⬇️ Download {selected_year} {doc_type} {volume}",
                    data=file,
                    file_name=f"{selected_year}_{doc_type}_{volume}.pdf",
                    mime="application/pdf"
                )
        except FileNotFoundError:
            st.sidebar.error("Requested file not found")
        except Exception as e:
            st.sidebar.error(f"Error downloading file: {str(e)}")
    if 'show_price_options' not in st.session_state:
        st.session_state.show_price_options = False
    # Add PRICE Rates download button
    # PRICE Rates button
    if st.sidebar.button("Download PRICE Rates"):
        toggle_section('show_price_options')
        st.rerun()

    if st.session_state.get('show_price_options', False):
        try:
            with open("Side Bar Files/PRICE Rates (DSR 21).xlsx", "rb") as file:
                st.sidebar.download_button(
                    label="⬇️ Download PRICE Rates (DSR 21) Excel",
                    data=file,
                    file_name="PRICE Rates (DSR 21).xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except FileNotFoundError:
            st.sidebar.error("PRICE Rates file not found")
        except Exception as e:
            st.sidebar.error(f"Error downloading PRICE Rates: {str(e)}")
            
    if 'show_dsr21basicrates_options' not in st.session_state:
        st.session_state.show_dsr21basicrates_options = False
    # Add DSR 21 Basic Rates download button
    # Basic Rates button
    if st.sidebar.button("Download Basic Rates"):
        toggle_section('show_dsr21basicrates_options')
        st.rerun()

    
    if st.session_state.get('show_dsr21basicrates_options', False):
        try:
            with open("Side Bar Files/DSR 21 Basic Rates.xlsx", "rb") as file:
                st.sidebar.download_button(
                    label="⬇️ Download Basic Rates (DSR 21) Excel",
                    data=file,
                    file_name="DSR 21 Basic Rates.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except FileNotFoundError:
            st.sidebar.error("DSR 21 Basic Rates file not found")
        except Exception as e:
            st.sidebar.error(f"Error downloading DSR 21 Basic Rates: {str(e)}")
        try:
            with open("Side Bar Files/DSR 21 Basic Rates.pdf", "rb") as file:
                st.sidebar.download_button(
                    label="⬇️ Download Basic Rates (DSR 21) PDF",
                    data=file,
                    file_name="DSR 21 Basic Rates.pdf",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except FileNotFoundError:
            st.sidebar.error("DSR 21 Basic Rates file not found")
        except Exception as e:
            st.sidebar.error(f"Error downloading DSR 21 Basic Rates: {str(e)}")    
    if 'show_priceapprovedmr_options' not in st.session_state:
        st.session_state.show_priceapprovedmr_options = False
    
    # Add PRICE Approved MR download button
    # PRICE Approved MR button
    if st.sidebar.button("PRICE Approved MR"):
        toggle_section('show_priceapprovedmr_options')
        st.rerun()
    
    if st.session_state.get('show_priceapprovedmr_options', False):
        try:
            with open("Side Bar Files/PRICE Approved MR.pdf", "rb") as file:
                st.sidebar.download_button(
                    label="⬇️ Download PRICE Approved MR PDF",
                    data=file,
                    file_name="PRICE Approved MR.pdf",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except FileNotFoundError:
            st.sidebar.error("PRICE Approved MR file not found")
        except Exception as e:
            st.sidebar.error(f"Error downloading PRICE Approved MR: {str(e)}")
    if 'show_gwd_options' not in st.session_state:
        st.session_state.show_gwd_options = False
    if 'show_costindex_options' not in st.session_state:
        st.session_state.show_costindex_options = False
    
    # Add Cost Index 2021 download button
    # Cost Index 2021 button
    if st.sidebar.button("Cost Index 2021"):
        toggle_section('show_costindex_options')
        st.rerun()
    
    if st.session_state.get('show_costindex_options', False):
        try:
            with open("Side Bar Files/Cost Index 2021.pdf", "rb") as file:
                st.sidebar.download_button(
                    label="⬇️ Download Cost Index 2021 PDF",
                    data=file,
                    file_name="Cost Index 2021.pdf",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except FileNotFoundError:
            st.sidebar.error("Cost Index 2021 file not found")
        except Exception as e:
            st.sidebar.error(f"Error downloading Cost Index 2021: {str(e)}")
    if 'show_gwd_options' not in st.session_state:
        st.session_state.show_gwd_options = False
        
    # Add GWD Data download button - similar to DSR download
    # GWD Data button
    if st.sidebar.button("Download GWD Data"):
        toggle_section('show_gwd_options')
        st.rerun()

    if st.session_state.get('show_gwd_options', False):
        try:
            # List files in the GWD Data directory
            import os
            gwd_files = []
            gwd_dir = "Side Bar Files/GWD Data"
            
            if os.path.exists(gwd_dir) and os.path.isdir(gwd_dir):
                gwd_files = [f for f in os.listdir(gwd_dir) if os.path.isfile(os.path.join(gwd_dir, f))]
            
            if not gwd_files:
                st.sidebar.warning("No files found in GWD Data directory")
            else:
                # Sort files alphabetically
                gwd_files.sort()
                
                # Create dropdown to select file
                selected_file = st.sidebar.selectbox(
                    "Select GWD Data File",
                    gwd_files,
                    key="gwd_file_select"
                )
                
                # Create download button for selected file
                file_path = os.path.join(gwd_dir, selected_file)
                
                # Determine MIME type based on file extension
                file_ext = os.path.splitext(selected_file)[1].lower()
                mime_types = {
                    '.pdf': 'application/pdf',
                    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    '.xls': 'application/vnd.ms-excel',
                    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    '.doc': 'application/msword',
                    '.txt': 'text/plain',
                    '.csv': 'text/csv'
                }
                mime_type = mime_types.get(file_ext, 'application/octet-stream')
                
                with open(file_path, "rb") as file:
                    st.sidebar.download_button(
                        label=f"⬇️ Download {selected_file}",
                        data=file,
                        file_name=selected_file,
                        mime=mime_type
                    )
                    
        except Exception as e:
            st.sidebar.error(f"Error accessing GWD Data: {str(e)}")    
    
    if 'show_templates_options' not in st.session_state:
        st.session_state.show_templates_options = False
    
    # Add Templates download button
    if st.sidebar.button("Download Templates"):
        toggle_section('show_templates_options')
        st.rerun()
    
    if st.session_state.get('show_templates_options', False):
        try:
            # List files in the Templates directory
            import os
            template_files = []
            templates_dir = "Data Base/Global Templates"
            
            if os.path.exists(templates_dir) and os.path.isdir(templates_dir):
                template_files = [f for f in os.listdir(templates_dir) if os.path.isfile(os.path.join(templates_dir, f))]
            
            if not template_files:
                st.sidebar.warning("No files found in Templates directory")
            else:
                # Sort files alphabetically
                template_files.sort()
                
                # Create dropdown to select file
                selected_file = st.sidebar.selectbox(
                    "Select Template File",
                    template_files,
                    key="template_file_select"
                )
                
                # Create download button for selected file
                file_path = os.path.join(templates_dir, selected_file)
                
                # Determine MIME type based on file extension
                file_ext = os.path.splitext(selected_file)[1].lower()
                mime_types = {
                    '.pdf': 'application/pdf',
                    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    '.xls': 'application/vnd.ms-excel',
                    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    '.doc': 'application/msword',
                    '.txt': 'text/plain',
                    '.csv': 'text/csv'
                }
                mime_type = mime_types.get(file_ext, 'application/octet-stream')
                
                with open(file_path, "rb") as file:
                    st.sidebar.download_button(
                        label=f"⬇️ Download {selected_file}",
                        data=file,
                        file_name=selected_file,
                        mime=mime_type
                    )
                    
        except Exception as e:
            st.sidebar.error(f"Error accessing Templates: {str(e)}")
    # Add to your session state initialization (if not already present)
    if 'show_pump_selector' not in st.session_state:
        st.session_state.show_pump_selector = False
    
    # In your sidebar section:
    # Pump Selector button
    if st.sidebar.button("Pump Selector"):
        toggle_section('show_pump_selector')
        st.rerun()
    
    if st.session_state.show_pump_selector:
        # In your Pump Selector section, change the button HTML to:
        st.sidebar.markdown("""
        <div style="background-color:#f0f2f6; padding:10px; border-radius:5px; margin-top:10px;">
            <p style="margin-bottom:10px;">Pump Selector will open in a new tab</p>
            <a href="https://gwdpumpdesign.streamlit.app/" target="_blank" class="pump-selector-btn" style="text-decoration:none;">
                <button style="background-color:#4CAF50; color:white; border:none; padding:8px 16px; 
                            text-align:center; display:inline-block; font-size:14px; margin:4px 2px; 
                            cursor:pointer; border-radius:4px;">
                    Open Pump Selector
                </button>
            </a>
            <p style="font-size:12px; color:#666; margin-top:10px;">
                If blocked, right-click → "Open in new tab"
            </p>
        </div>
        """, unsafe_allow_html=True)
    # Add logout button if authenticated  
    if st.sidebar.button("Logout"):
        st.session_state.authenticated = False
        st.session_state.logged_in_username = None
        st.rerun()        
